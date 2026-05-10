"""
Herramientas avanzadas de infraestructura: validación CMDB y editor de inventario Ansible.

Separado de infrastructure.py para mantener ese archivo por debajo de 1000 líneas.
"""

import csv
import io
import re
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone as dj_timezone

from ..models import (
    Company, LBPhysical, LBGuest, CMDBFieldConfig, AnsibleInventoryFile,
)
from .utils import _CMDB_COLS, _RESULT_COLS, _OK, _NA, _cmdb_compare_row


@login_required
@permission_required('lb_manager.view_lbphysical', raise_exception=True)
def cmdb_vs_lb_inventory(request):
    """
    Valida un CSV exportado del CMDB contra los registros LBPhysical + LBGuest de una empresa.

    GET  → muestra el formulario (selector de empresa + subida de archivo).
    POST → procesa el CSV y devuelve un reporte Excel con las diferencias.
    """
    companies = Company.objects.order_by('name')

    if request.method != 'POST':
        return render(request, 'lb_manager/cmdb_validation.html', {
            'companies': companies,
        })

    # ── read inputs ──────────────────────────────────────────────────────────
    company_id  = request.POST.get('company_id')
    upload_file = request.FILES.get('cmdb_file')

    if not company_id or not upload_file:
        messages.error(request, 'Debes seleccionar una empresa y subir el archivo CSV.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    _MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
    if upload_file.size > _MAX_UPLOAD_BYTES:
        messages.error(request, f'El archivo supera el límite de 10 MB ({upload_file.size // 1024 // 1024} MB).')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    # ── validate file type (extension + content-type + binary sniff) ──────────
    _ALLOWED_CSV_CONTENT_TYPES = {'text/csv', 'text/plain', 'application/csv', 'application/octet-stream'}
    reported_ct = upload_file.content_type.split(';')[0].strip().lower()
    filename_ok = upload_file.name.lower().endswith('.csv')
    ct_ok       = reported_ct in _ALLOWED_CSV_CONTENT_TYPES

    # Read first 512 bytes to detect binary content regardless of declared type
    header_bytes = upload_file.read(512)
    upload_file.seek(0)
    try:
        header_bytes.decode('utf-8-sig')
        is_text = True
    except UnicodeDecodeError:
        is_text = False

    if not filename_ok or not is_text:
        messages.error(request, 'Solo se aceptan archivos CSV. El archivo subido no es un CSV válido.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})
    if not ct_ok:
        messages.error(request, f'Tipo MIME no permitido: {reported_ct}. Se esperaba text/csv.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    try:
        company_obj = Company.objects.get(pk=company_id)
    except Company.DoesNotExist:
        messages.error(request, 'Empresa no encontrada.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    # ── parse CSV ────────────────────────────────────────────────────────────
    try:
        text      = upload_file.read().decode('utf-8-sig')  # handle BOM
        reader    = csv.DictReader(io.StringIO(text))
        cmdb_rows = list(reader)
    except (UnicodeDecodeError, csv.Error, LookupError, OSError) as exc:
        messages.error(request, f'Error al leer el archivo CSV: {exc}')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    if not cmdb_rows:
        messages.error(request, 'El archivo CSV está vacío.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    # ── load DB devices for the selected company ──────────────────────────────
    all_physicals = LBPhysical.objects.filter(company=company_obj).select_related('company')
    all_guests    = LBGuest.objects.filter(company=company_obj).select_related('company')

    db_by_name: dict[str, object] = {}
    for d in list(all_physicals) + list(all_guests):
        short = d.device.split('.')[0].lower()
        db_by_name[short] = d

    db_names = set(db_by_name.keys())

    # ── load configured expected values ──────────────────────────────────────
    ag_values = [
        v.lower() for v in
        CMDBFieldConfig.objects.filter(
            field_name=CMDBFieldConfig.Field.ASSIGNMENT_GROUP, active=True
        ).values_list('expected_value', flat=True)
    ]
    sg_values = [
        v.lower() for v in
        CMDBFieldConfig.objects.filter(
            field_name=CMDBFieldConfig.Field.SUPPORT_GROUP, active=True
        ).values_list('expected_value', flat=True)
    ]

    today = dj_timezone.localdate()

    # ── compare each CMDB row ─────────────────────────────────────────────────
    output_rows: list[dict] = []
    cmdb_names:  set[str]   = set()

    for row in cmdb_rows:
        cmdb_name = (row.get('name') or '').strip().lower()
        if cmdb_name:
            cmdb_names.add(cmdb_name)

        lb_device = db_by_name.get(cmdb_name)

        if not cmdb_name or lb_device is None:
            result = {
                'r_encontrado':       'No encontrado en DB' if cmdb_name else 'name vacío',
                'r_fqdn':             _NA,
                'r_asset_tag':        _NA,
                'r_company':          _NA,
                'r_used_for':         _NA,
                'r_ip_address':       _NA,
                'r_adm_ip_address':   _NA,
                'r_serial':           _NA,
                'r_last_discovered':  _NA,
                'r_assignment_group': _NA,
                'r_support_group':    _NA,
                'r_resumen':          'No encontrado en DB',
            }
        else:
            result = _cmdb_compare_row(row, lb_device, ag_values, sg_values, today)
            result['r_encontrado'] = 'Encontrado'

        output_rows.append({**row, **result})

    # ── devices in DB but NOT in CMDB ────────────────────────────────────────
    only_in_db = db_names - cmdb_names
    for short_name in sorted(only_in_db):
        lb = db_by_name[short_name]
        empty_cmdb = {col: '' for col in _CMDB_COLS}
        empty_cmdb['name'] = short_name
        empty_cmdb['fqdn'] = lb.device
        result = {col: _NA for col in _RESULT_COLS}
        result['r_encontrado'] = 'Solo en DB — no está en el archivo CMDB'
        result['r_resumen']    = 'Solo en DB'
        output_rows.append({**empty_cmdb, **result})

    # ── generate Excel with openpyxl ─────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN  = PatternFill('solid', fgColor='C6EFCE')
    RED    = PatternFill('solid', fgColor='FFC7CE')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')
    GREY   = PatternFill('solid', fgColor='D9D9D9')
    BLUE   = PatternFill('solid', fgColor='BDD7EE')
    HEADER_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Validación CMDB vs DB'

    cmdb_col_names = list(cmdb_rows[0].keys()) if cmdb_rows else _CMDB_COLS
    all_cols = list(cmdb_col_names) + _RESULT_COLS
    headers  = {col: idx + 1 for idx, col in enumerate(all_cols)}

    for col_name, col_idx in headers.items():
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill      = BLUE if col_name in _RESULT_COLS else GREY
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    for row_idx, row in enumerate(output_rows, start=2):
        for col_name, col_idx in headers.items():
            value = row.get(col_name, '')
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=False)
            cell.border    = THIN_BORDER
            if col_name in _RESULT_COLS:
                if value == _OK:
                    cell.fill = GREEN
                elif value == _NA:
                    cell.fill = GREY
                elif 'Revisar' in value:
                    cell.fill = YELLOW
                elif col_name == 'r_resumen' and value == 'Todo OK':
                    cell.fill = GREEN
                else:
                    cell.fill = RED

    for col_name, col_idx in headers.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 30 if col_name in _RESULT_COLS else 22

    ws2 = wb.create_sheet('Resumen')
    total_rows = len(output_rows)
    found      = sum(1 for r in output_rows if r.get('r_encontrado') == 'Encontrado')
    not_found  = sum(1 for r in output_rows if 'No encontrado' in r.get('r_encontrado', ''))
    only_db    = sum(1 for r in output_rows if 'Solo en DB' in r.get('r_encontrado', ''))
    all_ok     = sum(1 for r in output_rows if r.get('r_resumen') == 'Todo OK')
    with_diffs = sum(1 for r in output_rows if r.get('r_encontrado') == 'Encontrado'
                     and r.get('r_resumen') != 'Todo OK')

    summary = [
        ('Empresa validada',           company_obj.name),
        ('Fecha de validación',        str(today)),
        ('',                           ''),
        ('Total filas procesadas',     total_rows),
        ('Encontrados en DB',          found),
        ('No encontrados en DB',       not_found),
        ('Solo en DB (no en CMDB)',    only_db),
        ('',                           ''),
        ('Sin diferencias (Todo OK)',  all_ok),
        ('Con diferencias',            with_diffs),
    ]
    for s_row_idx, (label, value) in enumerate(summary, start=1):
        ws2.cell(row=s_row_idx, column=1, value=label).font = Font(bold=bool(label))
        ws2.cell(row=s_row_idx, column=2, value=value)
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = re.sub(r'[^\w\-.]', '_', company_obj.name)
    filename  = f'cmdb_validacion_{safe_name}_{today}.xlsx'
    response  = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Ansible Inventory File Editor ─────────────────────────────────────────────

def _parse_ini_hostnames(content: str) -> set[str]:
    """
    Retorna el conjunto de hostnames definidos en un string de inventario Ansible INI.

    Omite secciones [group:children] (son nombres de grupo, no hosts),
    líneas de comentario (#) y cabeceras de sección ([…]).
    Toma el primer token de cada línea restante como hostname.
    """
    hostnames: set[str] = set()
    in_children = False
    for raw in content.splitlines():
        line = raw.strip()
        if not line or line.startswith('#'):
            continue
        if line.startswith('['):
            in_children = ':children' in line
            continue
        if not in_children:
            hostnames.add(line.split()[0])
    return hostnames


# Prefijos del sistema de archivos que nunca deben leerse ni escribirse.
# La validación es una defensa en profundidad: file_path lo configura un admin,
# pero un error de configuración no debe exponer archivos críticos del SO.
_ANSIBLE_BLOCKED_PREFIXES: tuple[str, ...] = (
    '/etc/shadow', '/etc/gshadow', '/etc/sudoers', '/etc/passwd',
    '/proc/', '/sys/', '/dev/', '/root/',
)


def _resolve_ansible_path(raw_path: str) -> Path:
    """
    Resuelve y valida la ruta de un AnsibleInventoryFile.

    Normaliza la ruta (elimina '..' y symlinks) y comprueba que no apunte
    a ubicaciones críticas del sistema operativo.

    Args:
        raw_path: Valor del campo file_path almacenado en BD.

    Returns:
        Path resuelto y validado.

    Raises:
        ValueError: Si la ruta está bloqueada por política de seguridad.
    """
    resolved = Path(raw_path).resolve()
    for blocked in _ANSIBLE_BLOCKED_PREFIXES:
        if str(resolved).startswith(blocked):
            raise ValueError(f'Ruta bloqueada por política de seguridad: {resolved}')
    return resolved


@login_required
@permission_required('lb_manager.view_lbguest', raise_exception=True)
def ansible_inventory(request):
    """
    Editor de archivos de inventario Ansible.

    Muestra la lista de AnsibleInventoryFile configurados.
    Con uno seleccionado:
      GET  ?inv=<pk>  → lee el archivo del disco y muestra su contenido en un textarea,
                        listando los hosts de LBGuest para ese environment que aún no están.
      POST ?inv=<pk>  → escribe el contenido del textarea de vuelta al archivo.
    """
    inv_files     = AnsibleInventoryFile.objects.order_by('environment', 'name')
    inv_id        = request.GET.get('inv') or request.POST.get('inv')
    selected      = None
    file_content  = ''
    file_error    = None
    missing_hosts: list[str]  = []
    orphan_hosts:  list[dict] = []

    if inv_id:
        try:
            selected = AnsibleInventoryFile.objects.get(pk=inv_id)
        except AnsibleInventoryFile.DoesNotExist:
            messages.error(request, 'Inventario no encontrado.')

    if selected:
        if request.method == 'POST':
            if not request.user.has_perm('lb_manager.change_lbguest'):
                messages.error(request, 'No tienes permiso para modificar archivos de inventario.')
                file_content = request.POST.get('content', '')
            else:
                file_content = request.POST.get('content', '')
                try:
                    safe_path = _resolve_ansible_path(selected.file_path)
                    with open(safe_path, 'w', encoding='utf-8') as fh:
                        fh.write(file_content)
                    messages.success(request, f'Inventario guardado: {safe_path}')
                except ValueError as exc:
                    messages.error(request, str(exc))
                except OSError as exc:
                    messages.error(request, f'Error al guardar el archivo: {exc}')
        else:
            try:
                safe_path = _resolve_ansible_path(selected.file_path)
                with open(safe_path, 'r', encoding='utf-8') as fh:
                    file_content = fh.read()
            except ValueError as exc:
                file_error = str(exc)
            except FileNotFoundError:
                file_error = f'Archivo no encontrado: {selected.file_path}'
            except OSError as exc:
                file_error = f'Error al leer el archivo: {exc}'

        if file_content:
            inventory_hosts = _parse_ini_hostnames(file_content)

            db_hosts = list(
                LBGuest.objects
                .filter(environment__iexact=selected.environment)
                .values_list('device', flat=True)
                .order_by('device')
            )
            missing_hosts = [h for h in db_hosts if h not in inventory_hosts]

            db_all = {
                row['device']: row['environment']
                for row in LBGuest.objects.values('device', 'environment')
            }

            def _orphan_reason(h: str) -> str | None:
                if h not in db_all:
                    return 'No existe en la DB'
                if db_all[h].upper() != selected.environment.upper():
                    return f'En DB como {db_all[h]}'
                return None

            orphan_hosts = [
                {'host': h, 'reason': r}
                for h in sorted(inventory_hosts)
                if (r := _orphan_reason(h)) is not None
            ]

    return render(request, 'lb_manager/ansible_inventory.html', {
        'inv_files':     inv_files,
        'selected':      selected,
        'file_content':  file_content,
        'file_error':    file_error,
        'missing_hosts': missing_hosts,
        'orphan_hosts':  orphan_hosts,
    })
