"""
Shared helpers used across lb_manager view modules.
No view functions live here — only utilities.
"""

import logging
from collections import namedtuple
from datetime import timedelta
from typing import TYPE_CHECKING

from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.db.models import Count, Avg, ExpressionWrapper, F, DurationField, Q, QuerySet
from django.db.models.functions import TruncDate

from ..models import (
    LBGuest, LBPhysical, BitacoraHealth,
    ClientSSLProfile, SSLCert, VIP, LBDeviceChangeLog, SiteSettings,
)

if TYPE_CHECKING:
    from django.http import HttpRequest

logger = logging.getLogger(__name__)


def format_user_display(user):
    """Return 'First Last' or username, or '-' if user is None."""
    if not user:
        return '-'
    name = f"{user.first_name} {user.last_name}".strip()
    return name or user.username

# Ordered list of environments used across Wiki and other views
WIKI_ENVIRONMENTS = ['DRP', 'PRE-PRODUCTION', 'PRODUCTION']

# Environments considered "active" for compliance dashboards
# (DRP devices are excluded from operational compliance metrics)
COMPLIANCE_ENVIRONMENTS = ['PRE-PRODUCTION', 'PRODUCTION']

# LBPhysical.purpose values that count as load-balancers for inventory metrics
# (excludes 'Virtualizador' which only hosts guests). LBGuest has no purpose
# field and is implicitly counted as LTM.
LTM_GTM_PURPOSES = ['LTM', 'GTM']

_SITE_SETTINGS_CACHE_KEY = 'site_settings'
_SITE_SETTINGS_CACHE_TTL = 300  # 5 minutes


def get_site_settings():
    """
    Return the active SiteSettings instance, cached for 5 minutes.

    Using DatabaseCache, this value is shared across all Gunicorn workers,
    so the DB is queried at most once per 5 minutes cluster-wide. Call
    ``cache.delete('site_settings')`` from the admin save hook if you need
    immediate invalidation after a settings change.
    """
    cfg = cache.get(_SITE_SETTINGS_CACHE_KEY)
    if cfg is None:
        cfg = SiteSettings.objects.first()
        cache.set(_SITE_SETTINGS_CACHE_KEY, cfg, _SITE_SETTINGS_CACHE_TTL)
    return cfg

# Shown when mgmt_ip is changed and integration flags are auto-reset
_IP_RESET_WARNING_MSG = (
    'La MGMT IP fue modificada — Monitoreo, CyberArk y CMDB-SNMP fueron reiniciados. '
    'Recuerda volver a validar las integraciones.'
)


def _dt_length(request, cfg=None):
    """
    Return the DataTable page length from the request, capped to SiteSettings limits.

    Passing an already-loaded cfg avoids a second cache hit in views that
    already called get_site_settings() (e.g. vip_decommission_data).
    """
    if cfg is None:
        cfg = get_site_settings()
    default_page = cfg.datatable_default_page_length if cfg else 25
    max_rows = cfg.datatable_max_rows if cfg else 5000
    length = int(request.GET.get('length', default_page))
    if length == -1 or length > max_rows:
        length = max_rows
    return length


def _ha_map() -> dict[str, str]:
    """
    Mapa bidireccional ``device → ha_partner`` resolviendo LBPhysical y LBGuest.

    Útil cuando una vista necesita expandir un lookup por HA pair sin repetir
    la lógica de inferencia. Ejemplos de uso:

      - ``servicio_data`` para mostrar disponibilidad del VIP del peer cuando
        el match estricto por device falla.
      - ``servicio_delete_no_vip`` para no eliminar Servicios cuyo VIP existe
        en el HA partner.
      - ``servicio_sync`` para indexar Servicios por su par cruzado.

    Returns
    -------
    dict[str, str]
        Mapa ``device → partner`` con ambas direcciones de la pareja, derivado
        del campo ``ha_pair`` de cada hostname tanto en LBPhysical como en LBGuest.
    """
    pairs: dict[str, str] = {}
    pairs.update(
        LBPhysical.objects.filter(ha_pair__isnull=False).exclude(ha_pair='')
        .values_list('device', 'ha_pair')
    )
    pairs.update(
        LBGuest.objects.filter(ha_pair__isnull=False).exclude(ha_pair='')
        .values_list('device', 'ha_pair')
    )
    pairs.update({v: k for k, v in list(pairs.items()) if v not in pairs})
    return pairs


def _ha_group(ltm_fqdn: str) -> list[str]:
    """
    Given a device hostname, return a list with that device and its HA partner
    (if one is configured).  The lookup checks both LBGuest and LBPhysical.

    Examples
    --------
    _ha_group('lb-guest-01')  → ['lb-guest-01', 'lb-guest-02']
    _ha_group('lb-guest-01')  → ['lb-guest-01']   # if no ha_pair set
    """
    if not ltm_fqdn:
        return []

    group = {ltm_fqdn}

    # Check LBGuest
    guest = LBGuest.objects.filter(device=ltm_fqdn).values_list('ha_pair', flat=True).first()
    if guest:
        group.add(guest)
    else:
        # Maybe ltm_fqdn IS the ha_pair of another guest
        peer = LBGuest.objects.filter(ha_pair=ltm_fqdn).values_list('device', flat=True).first()
        if peer:
            group.add(peer)

    # Check LBPhysical (only LTM purpose pairs matter for VIP data)
    phys = LBPhysical.objects.filter(device=ltm_fqdn).values_list('ha_pair', flat=True).first()
    if phys:
        group.add(phys)
    else:
        peer = LBPhysical.objects.filter(ha_pair=ltm_fqdn).values_list('device', flat=True).first()
        if peer:
            group.add(peer)

    return list(group)


def _log_wiki_change(user, obj, action_flag):
    """Write a Django LogEntry for a wiki device change made via the app's UI."""
    LogEntry.objects.create(
        user_id=user.pk,
        content_type_id=ContentType.objects.get_for_model(obj).pk,
        object_id=str(obj.pk),
        object_repr=str(obj)[:200],
        action_flag=action_flag,
        change_message='Modified via UI' if action_flag == CHANGE else '',
    )


_FIELD_LABELS = {
    'ci_id': 'CI ID', 'mgmt_ip': 'MGMT IP', 'version': 'Version',
    'distro': 'Distro', 'model': 'Model', 'serial': 'Serial',
    'snow_link': 'SNOW Link', 'environment': 'Environment',
    'service': 'Service Date', 'vendor_support': 'Vendor Support',
    'location': 'Location', 'purpose': 'Purpose',
    'monitoreo': 'Monitoreo', 'cyberark': 'CyberArk',
    'user_cyberark': 'CyberArk User', 'company': 'Company',
    'datacenter': 'Datacenter', 'ansible_group': 'Ansible Group',
}


def _record_device_changes(user, old_obj, new_obj, device_type, changed_fields):
    """
    Save one LBDeviceChangeLog row per changed field, storing old and new values.

    ``changed_fields`` comes from Django's form.changed_data so only fields the
    user actually edited are inspected.  FK fields (company, datacenter) are
    stringified so the history shows names instead of raw IDs.
    """
    records = []
    for field in changed_fields:
        old_val = str(getattr(old_obj, field, '') or '')
        new_val = str(getattr(new_obj, field, '') or '')
        if old_val != new_val:
            records.append(LBDeviceChangeLog(
                device=str(new_obj.device),
                device_type=device_type,
                user=user,
                field_name=_FIELD_LABELS.get(field, field),
                old_value=old_val[:500],
                new_value=new_val[:500],
            ))
    if records:
        LBDeviceChangeLog.objects.bulk_create(records)


def _bita_avg_resolution():
    """Return the average resolution time in hours for CLOSED Bitácora tickets, or None."""
    avg_dur = BitacoraHealth.objects.filter(
        status=BitacoraHealth.Status.CLOSED,
        closed_at__isnull=False,
    ).annotate(
        duration=ExpressionWrapper(F('closed_at') - F('created_at'), output_field=DurationField())
    ).aggregate(avg=Avg('duration'))['avg']
    return round(avg_dur.total_seconds() / 3600, 1) if avg_dur else None


def _bita_weekly_trend(today, days=7):
    """Return list of {date, count} dicts for the last `days` days (tickets created)."""
    start = today - timedelta(days=days - 1)
    rows = (
        BitacoraHealth.objects
        .filter(created_at__date__gte=start)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    daily_map = {row['day']: row['count'] for row in rows}
    return [
        {'date': (start + timedelta(days=i)).strftime('%d/%m'),
         'count': daily_map.get(start + timedelta(days=i), 0)}
        for i in range(days)
    ]


_VIP_SSL_CERT_COUNTS_CACHE_KEY = 'vip_ssl_cert_counts'
_VIP_SSL_CERT_COUNTS_CACHE_TTL = 900  # 15 minutes


def _vip_ssl_cert_counts():
    """
    Returns (expired_enabled, soon_enabled):
      - expired_enabled : enabled VIPs whose client-side SSL cert is already expired
      - soon_enabled    : enabled VIPs whose client-side SSL cert expires within 30 days
    Uses expiration_timestamp for accuracy. Result is cached 15 minutes.
    """
    cached = cache.get(_VIP_SSL_CERT_COUNTS_CACHE_KEY)
    if cached is not None:
        return cached

    from datetime import datetime, timezone as dt_timezone

    today_dt  = datetime.now(dt_timezone.utc)
    now_ts    = int(today_dt.timestamp())
    soon_ts   = int((today_dt + timedelta(days=30)).timestamp())

    # Step 1 — certs split by status
    expired_fps = {}   # ltm_fqdn -> set of full_paths
    soon_fps    = {}

    for cert in SSLCert.objects.exclude(
            expiration_timestamp__isnull=True
    ).exclude(full_path__isnull=True).exclude(full_path='').only(
        'expiration_timestamp', 'full_path', 'ltm_fqdn'
    ):
        ts = cert.expiration_timestamp
        if ts < now_ts:
            expired_fps.setdefault(cert.ltm_fqdn, set()).add(cert.full_path)
        elif ts <= soon_ts:
            soon_fps.setdefault(cert.ltm_fqdn, set()).add(cert.full_path)

    if not expired_fps and not soon_fps:
        result = (0, 0)
        cache.set(_VIP_SSL_CERT_COUNTS_CACHE_KEY, result, _VIP_SSL_CERT_COUNTS_CACHE_TTL)
        return result

    all_fps = set()
    for s in expired_fps.values():
        all_fps |= s
    for s in soon_fps.values():
        all_fps |= s

    # Step 2 — profiles linked to those certs
    expired_profile_keys = set()
    soon_profile_keys    = set()

    for p in ClientSSLProfile.objects.filter(
        certificate_file__in=all_fps
    ).only('name', 'certificate_file', 'ltm_fqdn'):
        fp_set_exp  = expired_fps.get(p.ltm_fqdn, set())
        fp_set_soon = soon_fps.get(p.ltm_fqdn, set())
        if p.certificate_file in fp_set_exp:
            expired_profile_keys.add((p.ltm_fqdn, p.name))
        if p.certificate_file in fp_set_soon:
            soon_profile_keys.add((p.ltm_fqdn, p.name))

    if not expired_profile_keys and not soon_profile_keys:
        result = (0, 0)
        cache.set(_VIP_SSL_CERT_COUNTS_CACHE_KEY, result, _VIP_SSL_CERT_COUNTS_CACHE_TTL)
        return result

    # Step 3 — count enabled VIPs that reference those profiles (client-side)
    expired_vip_ids = set()
    soon_vip_ids    = set()

    for vip in VIP.objects.filter(
        enabled__in=['yes', 'enabled']
    ).exclude(profiles__isnull=True).exclude(profiles='').only(
        'id', 'ltm_fqdn', 'profiles'
    ):
        profiles_data = vip.profiles if isinstance(vip.profiles, list) else []
        for p in profiles_data:
            if not isinstance(p, dict) or p.get('context') != 'client-side':
                continue
            pname = p.get('name', '')
            key   = (vip.ltm_fqdn, pname)
            if key in expired_profile_keys:
                expired_vip_ids.add(vip.id)
            if key in soon_profile_keys:
                soon_vip_ids.add(vip.id)

    result = (len(expired_vip_ids), len(soon_vip_ids))
    cache.set(_VIP_SSL_CERT_COUNTS_CACHE_KEY, result, _VIP_SSL_CERT_COUNTS_CACHE_TTL)
    return result


_CERT_DATE_FORMATS = [
    '%b %d %H:%M:%S %Y',    # Apr 11 01:24:55 2026
    '%b  %d %H:%M:%S %Y',   # Apr  1 01:24:55 2026  (single-digit day)
    '%Y-%m-%d',             # 2026-04-11
    '%d/%m/%Y',             # 11/04/2026
    '%m/%d/%Y',             # 04/11/2026
]


def _parse_cert_expiration(cert):
    """
    Return the expiration date (datetime.date) for an SSLCert, or None if unparseable.

    Prefers ``expiration_timestamp`` (epoch int) for accuracy; falls back to
    the free-text ``expiration_date`` field, stripping trailing timezone tokens
    (e.g. "GTM", "UTC") before parsing.
    """
    from datetime import datetime, timezone as dt_timezone

    if cert.expiration_timestamp:
        return datetime.fromtimestamp(cert.expiration_timestamp, tz=dt_timezone.utc).date()
    if not cert.expiration_date:
        return None
    raw = cert.expiration_date.strip()
    parts = raw.rsplit(None, 1)
    candidates = [raw]
    if len(parts) == 2 and parts[1].isalpha() and 2 <= len(parts[1]) <= 5:
        candidates.append(parts[0])
    for candidate in candidates:
        for fmt in _CERT_DATE_FORMATS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except (ValueError, TypeError):
                pass
    return None


def _unassigned_pools_qs(ltm_fqdn_filter=None):
    """
    Return a queryset of Pools whose full_path is not referenced as
    default_pool in any VIP for the same ltm_fqdn.

    NOT EXISTS (
        SELECT 1 FROM vips v
        WHERE v.ltm_fqdn    = pools.ltm_fqdn
        AND   v.default_pool = pools.full_path
    )
    """
    from django.db.models.expressions import RawSQL
    from django.db.models import BooleanField
    from ..models import Pool

    in_vip = RawSQL(
        """
        EXISTS (
            SELECT 1 FROM vips v
            WHERE v.ltm_fqdn     = pools.ltm_fqdn
            AND   v.default_pool = pools.full_path
        )
        """,
        [],
        output_field=BooleanField(),
    )
    qs = Pool.objects.annotate(in_vip=in_vip).filter(in_vip=False)
    if ltm_fqdn_filter:
        qs = qs.filter(ltm_fqdn=ltm_fqdn_filter)
    return qs


def _unassigned_nodes_qs(ltm_fqdn_filter=None):
    """
    Return a queryset of LTMNodes whose address does not appear in any Pool's
    members array for the same ltm_fqdn.

    Uses a PostgreSQL JSONB containment subquery:
      NOT EXISTS (
          SELECT 1 FROM pools p
          WHERE p.ltm_fqdn = ltm_nodes.ltm_fqdn
          AND   p.members @> jsonb_build_array(jsonb_build_object('address', ltm_nodes.address))
      )
    """
    from django.db.models.expressions import RawSQL
    from django.db.models import BooleanField
    from ..models import LTMNode

    in_pool = RawSQL(
        """
        EXISTS (
            SELECT 1 FROM pools p
            WHERE p.ltm_fqdn = ltm_nodes.ltm_fqdn
            AND   p.members @> jsonb_build_array(jsonb_build_object('address', ltm_nodes.address))
        )
        """,
        [],
        output_field=BooleanField(),
    )
    qs = LTMNode.objects.annotate(in_pool=in_pool).filter(in_pool=False)
    if ltm_fqdn_filter:
        qs = qs.filter(ltm_fqdn=ltm_fqdn_filter)
    return qs


# CMDB validation constants
_CMDB_COLS = [
    'name', 'asset_tag', 'fqdn', 'assignment_group', 'model_id', 'company',
    'u_delivery', 'user_for', 'ip_address', 'u_adm_ip_address',
    'serial_number', 'u_status', 'managed_by', 'sys_updated_on',
    'sys_update_by', 'last_discovered', 'support_group',
    'u_integration', 'u_distro_network',
]

# Result column names appended to the output file
_RESULT_COLS = [
    'r_encontrado',      # JOIN result: name vs device short hostname
    'r_fqdn',
    'r_asset_tag',       # asset_tag vs ci_id (validation, not join key)
    'r_company',
    'r_used_for',
    'r_ip_address',
    'r_adm_ip_address',
    'r_serial',
    'r_last_discovered',
    'r_assignment_group',
    'r_support_group',
    'r_resumen',
]

_OK  = 'OK'
_NA  = 'N/A'


def _cmdb_compare_row(cmdb_row, lb_device, ag_values, sg_values, today):
    """
    Compare one CMDB row against the matched LB device.
    Returns a dict of result columns.
    """
    from datetime import datetime

    res = {}

    # ── fqdn (CMDB) vs full device FQDN (DB) ─────────────────────────────────
    cmdb_fqdn = (cmdb_row.get('fqdn') or '').strip().lower()
    db_fqdn   = (lb_device.device or '').strip().lower()
    res['r_fqdn'] = (
        _OK if cmdb_fqdn == db_fqdn
        else f'Diferente (CMDB: "{cmdb_row.get("fqdn","")}" / DB: "{lb_device.device}")'
    )

    # ── asset_tag (CMDB) vs ci_id (DB) ───────────────────────────────────────
    cmdb_at = (cmdb_row.get('asset_tag') or '').strip().lower()
    db_ci   = (lb_device.ci_id or '').strip().lower()
    res['r_asset_tag'] = (
        _OK if cmdb_at == db_ci
        else f'Diferente (CMDB: "{cmdb_row.get("asset_tag","")}" / DB ci_id: "{lb_device.ci_id}")'
    )

    # ── company ──────────────────────────────────────────────────────────────
    cmdb_company = (cmdb_row.get('company') or '').strip().lower()
    db_company   = (lb_device.company.name if lb_device.company else '').lower()
    res['r_company'] = (
        _OK if cmdb_company == db_company
        else f'Diferente (CMDB: "{cmdb_row.get("company","")}" / DB: "{lb_device.company.name if lb_device.company else ""}")'
    )

    # ── used_for / environment ────────────────────────────────────────────────
    cmdb_env_raw = (cmdb_row.get('used_for') or cmdb_row.get('user_for') or '').strip()
    cmdb_env = cmdb_env_raw.lower()
    db_env   = (lb_device.environment or '').strip().lower()
    res['r_used_for'] = (
        _OK if cmdb_env == db_env
        else f'Diferente (CMDB: "{cmdb_env_raw}" / DB: "{lb_device.environment}")'
    )

    # ── ip_address vs mgmt_ip ─────────────────────────────────────────────────
    cmdb_ip = (cmdb_row.get('ip_address') or '').strip()
    db_ip   = (lb_device.mgmt_ip or '').strip()
    res['r_ip_address'] = (
        _OK if cmdb_ip == db_ip
        else f'Diferente (CMDB: "{cmdb_ip}" / DB: "{db_ip}")'
    )

    # ── u_adm_ip_address vs mgmt_ip ──────────────────────────────────────────
    cmdb_adm = (cmdb_row.get('u_adm_ip_address') or cmdb_row.get('u_admin_ip_address') or '').strip()
    res['r_adm_ip_address'] = (
        _OK if cmdb_adm == db_ip
        else f'Diferente (CMDB: "{cmdb_adm}" / DB: "{db_ip}")'
    )

    # ── serial_number vs serial ───────────────────────────────────────────────
    cmdb_serial = (cmdb_row.get('serial_number') or '').strip().lower()
    db_serial   = (lb_device.serial or '').strip().lower()
    res['r_serial'] = (
        _OK if cmdb_serial == db_serial
        else f'Diferente (CMDB: "{cmdb_row.get("serial_number","")}" / DB: "{lb_device.serial}")'
    )

    # ── last_discovered ───────────────────────────────────────────────────────
    raw_ld = (cmdb_row.get('last_discovered') or '').strip()
    if not raw_ld:
        res['r_last_discovered'] = 'Sin fecha'
    else:
        try:
            # Accept common formats: YYYY-MM-DD, YYYY-MM-DD HH:MM:SS, MM/DD/YYYY
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                try:
                    ld_date = datetime.strptime(raw_ld.split(' ', maxsplit=1)[0], fmt.split(' ', maxsplit=1)[0]).date()
                    break
                except ValueError:
                    continue
            else:
                ld_date = None

            if ld_date is None:
                res['r_last_discovered'] = f'Formato no reconocido: {raw_ld}'
            elif (today - ld_date).days > 30:
                res['r_last_discovered'] = f'Revisar — último descubrimiento hace {(today - ld_date).days} días ({raw_ld})'
            else:
                res['r_last_discovered'] = _OK
        except (ValueError, OverflowError, OSError):
            logger.exception('Error parsing last_discovered date: %r', raw_ld)
            res['r_last_discovered'] = f'Formato no reconocido: {raw_ld}'

    # ── assignment_group ──────────────────────────────────────────────────────
    cmdb_ag = (cmdb_row.get('assignment_group') or '').strip().lower()
    if not ag_values:
        res['r_assignment_group'] = _NA
    elif any(cmdb_ag == v for v in ag_values):
        res['r_assignment_group'] = _OK
    else:
        expected = ' / '.join(f'"{v}"' for v in ag_values)
        res['r_assignment_group'] = f'No coincide (CMDB: "{cmdb_row.get("assignment_group","")}" — esperado: {expected})'

    # ── support_group ─────────────────────────────────────────────────────────
    cmdb_sg = (cmdb_row.get('support_group') or '').strip().lower()
    if not sg_values:
        res['r_support_group'] = _NA
    elif any(cmdb_sg == v for v in sg_values):
        res['r_support_group'] = _OK
    else:
        expected = ' / '.join(f'"{v}"' for v in sg_values)
        res['r_support_group'] = f'No coincide (CMDB: "{cmdb_row.get("support_group","")}" — esperado: {expected})'

    # ── resumen ───────────────────────────────────────────────────────────────
    diffs = [k for k, v in res.items() if v not in (_OK, _NA) and not v.startswith('OK')]
    res['r_resumen'] = 'Todo OK' if not diffs else f'{len(diffs)} diferencia(s): {", ".join(diffs)}'

    return res


# ══════════════════════════════════════════════════════════════════════════════
# Generic reusable helpers (Phase 1 refactoring — used across multiple views)
# ══════════════════════════════════════════════════════════════════════════════

_DTParams = namedtuple('_DTParams', ['draw', 'start', 'length', 'search', 'col_idx', 'col_dir'])


def _dt_params(
    request: 'HttpRequest',
    cfg: object = None,
    *,
    default_col: int = 0,
    default_dir: str = 'asc',
) -> '_DTParams':
    """
    Extract and parse standard DataTables AJAX request parameters.

    Centralizes the parsing block repeated in every ``*_data(request)`` view.
    Returns a NamedTuple to support tuple unpacking in a single line:

        draw, start, length, search, col_idx, col_dir = _dt_params(request)

    Args:
        request: Incoming GET request from a DataTables AJAX call.
        cfg: Optional pre-loaded SiteSettings instance to forward to
             :func:`_dt_length` and avoid a second cache hit.
        default_col: Default value for ``order[0][column]`` when missing.
        default_dir: Default value for ``order[0][dir]`` when missing.

    Returns:
        NamedTuple with fields ``draw``, ``start``, ``length``, ``search``,
        ``col_idx`` and ``col_dir``.
    """
    return _DTParams(
        draw=int(request.GET.get('draw', 1)),
        start=int(request.GET.get('start', 0)),
        length=_dt_length(request, cfg),
        search=request.GET.get('search[value]', '').strip(),
        col_idx=int(request.GET.get('order[0][column]', default_col)),
        col_dir=request.GET.get('order[0][dir]', default_dir),
    )


def apply_search_filter(qs: QuerySet, term: str, fields: list[str]) -> QuerySet:
    """
    Apply a case-insensitive multi-field OR search to a queryset.

    Replaces the inline ``Q(field1__icontains=term) | Q(field2__icontains=term) | ...``
    block repeated across DataTables AJAX views.

    Args:
        qs: Source queryset to filter.
        term: Search term entered by the user. If empty, ``qs`` is returned
              unchanged so the caller does not need a separate ``if term``.
        fields: List of field names (supports Django lookup paths like
                ``company__name``) that will be OR-joined with ``__icontains``.

    Returns:
        Filtered queryset (or the original one if ``term`` is empty).
    """
    if not term or not fields:
        return qs
    q = Q()
    for field in fields:
        q |= Q(**{f'{field}__icontains': term})
    return qs.filter(q)


def _excel_headers(ws: object, headers: list[str]) -> None:
    """
    Apply the project's standard header styling to row 1 of a worksheet.

    Header background: ``#1F4E79`` (corporate blue), white bold font,
    centered. Used by every export view that builds an .xlsx file.

    Args:
        ws: ``openpyxl`` Worksheet object. Row 1 will be overwritten.
        headers: Column header labels to write left-to-right.
    """
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
