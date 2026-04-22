"""
VIP views: CRUD, lookup, export, data, dormant and historical snapshots.
Pool, Node, SelfIP, SNAT and Servicio views live in pools.py.
"""

import io
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import connection
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils import timezone as dj_timezone
from django.views.generic import ListView, UpdateView

from ..models import (
    VIP, Pool, LBVIPHistorical, SiteSettings, LBPhysical, LBGuest,
)
from ..forms import VIPForm
from .utils import _dt_length, _ha_group


class VIPUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Manual override of a VIP's operational state.

    VIPs are normally synced from F5 (read-only), but operators can correct
    enabled/availability values here.  Saving also writes a snapshot row to
    LBVIPHistorical so the change is reflected in the dormant-VIP report once
    the inactivity period is reached.
    """

    model = VIP
    form_class = VIPForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('vip_list')
    permission_required = 'lb_manager.change_vip'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit VIP: {self.object.name}'
        ctx['list_url'] = reverse_lazy('vip_list')
        ctx['list_label'] = 'VIPs'
        ctx['is_edit'] = True
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        vip = self.object
        # Insert a historical snapshot so the change is tracked over time
        LBVIPHistorical.objects.create(
            name=vip.name,
            full_path=vip.full_path,
            destination=vip.destination,
            destination_address=vip.destination_address,
            destination_port=vip.destination_port,
            protocol=vip.protocol,
            type=vip.type,
            enabled=vip.enabled,
            default_pool=vip.default_pool,
            snat_type=vip.snat_type,
            snat_pool=vip.snat_pool,
            source_address=vip.source_address,
            connection_limit=vip.connection_limit,
            persistence_profile=vip.persistence_profile,
            profiles=vip.profiles,
            policies=vip.policies,
            description=vip.description,
            ltm_fqdn=vip.ltm_fqdn,
            date=dj_timezone.now(),
        )
        messages.success(self.request, f'VIP "{vip.name}" actualizada correctamente.')
        return response


# ---------------------------------------------------------------------------
# VIP Lookup
# ---------------------------------------------------------------------------

@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def vip_lookup(request):
    """
    VIP search tool — find a VIP and see its full configuration in one view.

    Accepts a POST form with a list of search terms (one per line) and a
    search type:
      - name_exact    : Match VIP names exactly (case-sensitive list lookup)
      - name_contains : Partial name match (any term is a substring)
      - destination   : Match destination_address exactly (exact IP lookup)

    For each matching VIP the response also includes the associated pool's
    members (real servers) and health monitors, and parses the JSON ``profiles``
    field into a readable list of profile names.
    """
    names_raw = ''
    search_type = 'name_exact'
    results = []
    searched = False

    if request.method == 'POST':
        names_raw = request.POST.get('names', '')
        search_type = request.POST.get('search_type', 'name_exact')
        searched = True
        terms = [n.strip() for n in names_raw.splitlines() if n.strip()]

        if terms:
            if search_type == 'name_exact':
                vips = VIP.objects.filter(name__in=terms)
            elif search_type == 'name_contains':
                q = Q()
                for t in terms:
                    q |= Q(name__icontains=t)
                vips = VIP.objects.filter(q)
            elif search_type == 'destination':
                vips = VIP.objects.filter(destination_address__in=terms)
            else:
                vips = VIP.objects.none()

            # Build pool map: (full_path, ltm_fqdn) → Pool
            pool_map = {
                (p.full_path, p.ltm_fqdn): p
                for p in Pool.objects.filter(
                    full_path__in=vips.values_list('default_pool', flat=True)
                )
            }

            for vip in vips:
                pool = pool_map.get((vip.default_pool, vip.ltm_fqdn))

                # Parse profiles (JSONField — already deserialized by Django)
                try:
                    profiles_data = vip.profiles if vip.profiles else []
                    if isinstance(profiles_data, list):
                        profiles_names = [p.get('name', '') for p in profiles_data if isinstance(p, dict)]
                    else:
                        profiles_names = []
                except (ValueError, TypeError):
                    profiles_names = []

                members = []
                monitors = []
                if pool:
                    raw_members = pool.members or []
                    for m in raw_members:
                        if isinstance(m, dict):
                            # Port lives after the last ':' in full_path, e.g. /Common/node:443
                            full_path = m.get('full_path', '')
                            port = full_path.rsplit(':', 1)[-1] if ':' in full_path else ''
                            members.append({
                                'name': m.get('name', '-'),
                                'address': m.get('address', '-'),
                                'port': port,
                            })
                    raw_monitors = pool.monitors or []
                    monitors = raw_monitors if isinstance(raw_monitors, list) else []

                results.append({
                    'ltm_fqdn': vip.ltm_fqdn,
                    'name': vip.name,
                    'destination_address': vip.destination_address,
                    'destination_port': vip.destination_port,
                    'availability_status': vip.availability_status,
                    'type': vip.type,
                    'enabled': vip.enabled,
                    'profiles': profiles_names,
                    'default_pool': vip.default_pool,
                    'members': members,
                    'monitors': monitors,
                })

    return render(request, 'lb_manager/vip_lookup.html', {
        'names_raw': names_raw,
        'search_type': search_type,
        'results': results,
        'searched': searched,
    })


@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def vip_lookup_export(request):
    """
    Export the full VIP inventory as an Excel file.

    Includes HA Pair (from LBGuest), pool members and monitors for every VIP.
    Uses 3 queries: LBGuest, VIP, Pool — no N+1.
    """
    from openpyxl import Workbook  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Font, PatternFill, Alignment  # pylint: disable=import-outside-toplevel

    # 1. device → ha_pair
    ha_map = {g.device: g.ha_pair or '-' for g in LBGuest.objects.only('device', 'ha_pair')}

    # 2. All VIPs
    vips = list(VIP.objects.order_by('ltm_fqdn', 'name'))

    # 3. Pools referenced by any VIP
    pool_map = {
        (p.full_path, p.ltm_fqdn): p
        for p in Pool.objects.filter(
            full_path__in={v.default_pool for v in vips if v.default_pool}
        )
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'VIPs'

    headers = [
        'LTM', 'HA Pair', 'VIP Name', 'Destination IP', 'Port',
        'Enabled', 'Availability', 'Default Pool', 'Pool Members', 'Monitors',
    ]
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for vip in vips:
        pool = pool_map.get((vip.default_pool, vip.ltm_fqdn))
        members_str = '-'
        monitors_str = '-'
        if pool:
            raw = pool.members or []
            parts = []
            for m in raw:
                if isinstance(m, dict):
                    fp = m.get('full_path', '')
                    port = fp.rsplit(':', 1)[-1] if ':' in fp else ''
                    addr = m.get('address', '-')
                    name = m.get('name', '-')
                    parts.append(f'{name} ({addr}{":" + port if port else ""})')
            members_str = ' | '.join(parts) if parts else '-'
            raw_mon = pool.monitors or []
            monitors_str = ', '.join(raw_mon) if isinstance(raw_mon, list) else '-'

        ws.append([
            vip.ltm_fqdn or '-',
            ha_map.get(vip.ltm_fqdn, '-'),
            vip.name or '-',
            vip.destination_address or '-',
            vip.destination_port or '-',
            vip.enabled or '-',
            vip.availability_status or '-',
            vip.default_pool or '-',
            members_str,
            monitors_str,
        ])

    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="vip_inventory.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# VIP List + AJAX
# ---------------------------------------------------------------------------

class VIPListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Renders the VIP list page shell. Rows are loaded via vip_data."""

    model = VIP
    template_name = 'lb_manager/vip_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_vip'
    raise_exception = True

    def get_queryset(self):
        return VIP.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Virtual IPs (VIPs)'
        ctx['page_icon'] = 'fa-globe'
        ctx['total'] = VIP.objects.count()
        return ctx


@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def vip_data(request):
    """
    AJAX endpoint — returns all 35 VIP columns as JSON for DataTables.

    Columns (0-indexed):
      0  id                          1  name
      2  full_path                   3  destination
      4  destination_address         5  destination_port
      6  protocol                    7  type
      8  enabled                     9  availability_status
      10 status_reason               11 default_pool
      12 snat_type                   13 snat_pool
      14 source_address              15 source_port_behavior
      16 persistence_profile         17 profiles
      18 policies                    19 description
      20 connection_limit            21 connection_mirror_enabled
      22 rate_limit                  23 rate_limit_mode
      24 rate_limit_destination_mask 25 translate_address
      26 translate_port              27 nat64_enabled
      28 cmp_enabled                 29 cmp_mode
      30 hardware_syn_cookie_instances 31 syn_cookies_status
      32 auto_lasthop                33 gtm_score
      34 ltm_fqdn                   35 ha_pair
    """
    COLUMNS = [
        'id', 'name', 'full_path', 'destination', 'destination_address',
        'destination_port', 'protocol', 'type', 'enabled', 'availability_status',
        'status_reason', 'default_pool', 'snat_type', 'snat_pool', 'source_address',
        'source_port_behavior', 'persistence_profile',
        'profiles', 'profiles', 'profiles',           # client / server / all (same DB col, split in Python)
        'policies', 'description', 'connection_limit',
        'connection_mirror_enabled', 'rate_limit', 'rate_limit_mode',
        'rate_limit_destination_mask', 'translate_address', 'translate_port',
        'nat64_enabled', 'cmp_enabled', 'cmp_mode', 'hardware_syn_cookie_instances',
        'syn_cookies_status', 'auto_lasthop', 'gtm_score', 'ltm_fqdn',
    ]

    draw    = int(request.GET.get('draw', 1))
    start   = int(request.GET.get('start', 0))
    length  = _dt_length(request)
    search  = request.GET.get('search[value]', '').strip()
    col_idx = int(request.GET.get('order[0][column]', 1))
    col_dir = request.GET.get('order[0][dir]', 'asc')

    qs = VIP.objects.all()
    total = qs.count()

    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(destination_address__icontains=search) |
            Q(destination_port__icontains=search) |
            Q(protocol__icontains=search) |
            Q(type__icontains=search) |
            Q(availability_status__icontains=search) |
            Q(default_pool__icontains=search) |
            Q(snat_type__icontains=search) |
            Q(source_address__icontains=search) |
            Q(profiles__icontains=search) |
            Q(description__icontains=search) |
            Q(policies__icontains=search) |
            Q(ltm_fqdn__icontains=search)
        )
    filtered = qs.count()

    order_col = COLUMNS[col_idx] if col_idx < len(COLUMNS) else 'name'
    if col_dir == 'desc':
        order_col = f'-{order_col}'
    qs = qs.order_by(order_col)[start:start + length]

    def s(val):
        return val if val is not None else '-'

    def parse_profiles(raw):
        """Return (client_names, server_names, all_names) parsed from the profiles JSON."""
        client, server, all_ = [], [], []
        if not raw:
            return '-', '-', '-'
        try:
            items = json.loads(raw) if isinstance(raw, str) else raw
            if not isinstance(items, list):
                return '-', '-', '-'
            for item in items:
                ctx  = (item.get('context') or '').lower()
                name = item.get('name') or item.get('full_path') or ''
                if ctx == 'client-side':
                    client.append(name)
                elif ctx == 'server-side':
                    server.append(name)
                elif ctx == 'all':
                    all_.append(name)
        except (json.JSONDecodeError, TypeError, AttributeError):
            return '-', '-', '-'
        return (', '.join(client) or '-',
                ', '.join(server) or '-',
                ', '.join(all_)   or '-')

    ha_lookup = dict(LBPhysical.objects.values_list('device', 'ha_pair'))
    ha_lookup.update(LBGuest.objects.values_list('device', 'ha_pair'))

    data = []
    for v in qs:
        prof_client, prof_server, prof_all = parse_profiles(v.profiles)
        data.append([
            v.id,                           # 0
            s(v.name),                      # 1
            s(v.full_path),                 # 2
            s(v.destination),               # 3
            s(v.destination_address),       # 4
            s(v.destination_port),          # 5
            s(v.protocol),                  # 6
            s(v.type),                      # 7
            s(v.enabled),                   # 8
            s(v.availability_status),       # 9
            s(v.status_reason),             # 10
            s(v.default_pool),              # 11
            s(v.snat_type),                 # 12
            s(v.snat_pool),                 # 13
            s(v.source_address),            # 14
            s(v.source_port_behavior),      # 15
            s(v.persistence_profile),       # 16
            prof_client,                    # 17
            prof_server,                    # 18
            prof_all,                       # 19
            s(v.policies),                  # 20
            s(v.description),               # 21
            s(v.connection_limit),          # 22
            s(v.connection_mirror_enabled), # 23
            s(v.rate_limit),                # 24
            s(v.rate_limit_mode),           # 25
            s(v.rate_limit_destination_mask), # 26
            s(v.translate_address),         # 27
            s(v.translate_port),            # 28
            s(v.nat64_enabled),             # 29
            s(v.cmp_enabled),               # 30
            s(v.cmp_mode),                  # 31
            s(v.hardware_syn_cookie_instances), # 32
            s(v.syn_cookies_status),        # 33
            s(v.auto_lasthop),              # 34
            s(v.gtm_score),                 # 35
            s(v.ltm_fqdn),                  # 36
            ha_lookup.get(v.ltm_fqdn) or '-',  # 37 ha_pair
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': filtered,
        'data': data,
    })


# ---------------------------------------------------------------------------
# VIP Dormant — VIPs that have been inactive for 3+ months
# ---------------------------------------------------------------------------

class VIPDormantListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    VIPs that have been consistently inactive (enabled='no' OR
    availability not 'available') for 3 or more months, identified by
    cross-referencing current VIP state with historical snapshots.
    """

    model = VIP
    template_name = 'lb_manager/vip_dormant_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_vip'
    raise_exception = True

    def get_queryset(self):
        return VIP.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'VIPs Inactivas (3+ meses)'
        ctx['page_icon'] = 'fa-circle-xmark'
        ctx['ltm_fqdns'] = (
            VIP.objects.exclude(ltm_fqdn__isnull=True).exclude(ltm_fqdn='')
            .values_list('ltm_fqdn', flat=True).distinct().order_by('ltm_fqdn')
        )
        cfg = SiteSettings.objects.first()
        lookback = cfg.decommission_lookback_months if cfg else 3
        with connection.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*)
                FROM vips v
                JOIN (
                    SELECT name, ltm_fqdn
                    FROM lb_vips_historical
                    WHERE date >= NOW() - (INTERVAL '1 month' * %s)
                    GROUP BY name, ltm_fqdn
                    HAVING COUNT(*) > 0
                       AND COUNT(*) = COUNT(CASE WHEN LOWER(enabled) = 'no' THEN 1 END)
                ) h ON v.name = h.name AND v.ltm_fqdn = h.ltm_fqdn
            """, [lookback])
            ctx['total'] = cur.fetchone()[0]
        return ctx


@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def vip_dormant_data(request):
    """
    AJAX endpoint — returns VIPs dormant for 3+ months for DataTables.

    A VIP is considered dormant when ALL snapshots in lb_vips_historical for the
    last 3 months have enabled='no'. Enabled and availability_status come from
    the most recent snapshot in lb_vips_historical (not the vips table).

    Columns (0-indexed):
      0  name   1  ltm_fqdn   2  destination_address   3  destination_port
      4  enabled (latest hist)   5  availability_status (latest hist)
      6  default_pool   7  first_inactive   8  days_inactive   9  id
    """
    # Note: decommission_candidate removed — dormant status is computed dynamically
    cfg        = SiteSettings.objects.first()
    lookback   = cfg.decommission_lookback_months if cfg else 3
    draw       = int(request.GET.get('draw', 1))
    start      = int(request.GET.get('start', 0))
    length     = _dt_length(request, cfg)
    search     = request.GET.get('search[value]', '').strip()
    col_idx    = int(request.GET.get('order[0][column]', 8))
    col_dir    = request.GET.get('order[0][dir]', 'desc')
    ltm_filter = request.GET.get('ltm_fqdn', '').strip()

    ORDER_COLS = {
        0: 'v.name', 1: 'v.ltm_fqdn', 2: 'v.destination_address',
        3: 'v.destination_port', 4: 'lat.enabled', 5: 'lat.availability_status',
        6: 'v.default_pool', 7: 'fi.first_inactive', 8: 'days_inactive',
    }
    order_expr = ORDER_COLS.get(col_idx, 'days_inactive')
    order_dir  = 'DESC' if col_dir == 'desc' else 'ASC'

    extra_clauses = []
    params: list = []
    if search:
        extra_clauses.append("""
            (v.name ILIKE %s OR v.ltm_fqdn ILIKE %s
             OR v.destination_address ILIKE %s OR v.default_pool ILIKE %s)
        """)
        like = f'%{search}%'
        params += [like, like, like, like]
    if ltm_filter:
        ha_devs = _ha_group(ltm_filter)
        placeholders = ', '.join(['%s'] * len(ha_devs))
        extra_clauses.append(f'v.ltm_fqdn IN ({placeholders})')
        params += ha_devs

    search_clause = ('AND ' + ' AND '.join(extra_clauses)) if extra_clauses else ''

    base_sql = f"""
        SELECT
            v.name,
            v.ltm_fqdn,
            COALESCE(v.destination_address, '-')          AS destination_address,
            COALESCE(v.destination_port::text, '-')       AS destination_port,
            COALESCE(lat.enabled, '-')                    AS enabled,
            COALESCE(lat.availability_status, '-')        AS availability_status,
            COALESCE(v.default_pool, '-')                 AS default_pool,
            fi.first_inactive::date                       AS first_inactive,
            (CURRENT_DATE - fi.first_inactive::date)      AS days_inactive,
            v.id
        FROM vips v
        JOIN (
            -- ALL snapshots in the last 3 months must have enabled='no'
            SELECT name, ltm_fqdn
            FROM lb_vips_historical
            WHERE date >= NOW() - (INTERVAL '1 month' * %s)
            GROUP BY name, ltm_fqdn
            HAVING COUNT(*) > 0
               AND COUNT(*) = COUNT(CASE WHEN LOWER(enabled) = 'no' THEN 1 END)
        ) h ON v.name = h.name AND v.ltm_fqdn = h.ltm_fqdn
        JOIN (
            -- Most recent snapshot values (enabled, availability_status)
            SELECT DISTINCT ON (name, ltm_fqdn)
                name, ltm_fqdn, enabled, availability_status
            FROM lb_vips_historical
            ORDER BY name, ltm_fqdn, date DESC
        ) lat ON v.name = lat.name AND v.ltm_fqdn = lat.ltm_fqdn
        JOIN (
            -- Earliest ever inactive snapshot across all history (for display)
            SELECT name, ltm_fqdn, MIN(date) AS first_inactive
            FROM lb_vips_historical
            WHERE LOWER(enabled) = 'no'
            GROUP BY name, ltm_fqdn
        ) fi ON v.name = fi.name AND v.ltm_fqdn = fi.ltm_fqdn
        WHERE 1=1
        {search_clause}
    """

    base_params = [lookback] + params
    with connection.cursor() as cur:
        cur.execute(f'SELECT COUNT(*) FROM ({base_sql}) sub', base_params)
        total_filtered = cur.fetchone()[0]

        cur.execute(
            f'{base_sql} ORDER BY {order_expr} {order_dir} LIMIT %s OFFSET %s',
            base_params + [length, start],
        )
        rows = cur.fetchall()

    # Count for total (no search filter)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT COUNT(*)
            FROM vips v
            JOIN (
                SELECT name, ltm_fqdn
                FROM lb_vips_historical
                WHERE date >= NOW() - (INTERVAL '1 month' * %s)
                GROUP BY name, ltm_fqdn
                HAVING COUNT(*) > 0
                   AND COUNT(*) = COUNT(CASE WHEN LOWER(enabled) = 'no' THEN 1 END)
            ) h ON v.name = h.name AND v.ltm_fqdn = h.ltm_fqdn
        """, [lookback])
        total = cur.fetchone()[0]

    data = [
        [
            row[0],                                            # 0 name
            row[1],                                            # 1 ltm_fqdn
            row[2],                                            # 2 destination_address
            row[3],                                            # 3 destination_port
            row[4],                                            # 4 enabled (latest hist)
            row[5],                                            # 5 availability_status (latest hist)
            row[6],                                            # 6 default_pool
            row[7].strftime('%d/%m/%Y') if row[7] else '-',   # 7 first_inactive
            int(row[8]) if row[8] is not None else 0,         # 8 days_inactive
            row[9],                                            # 9 id
        ]
        for row in rows
    ]

    return JsonResponse({'draw': draw, 'recordsTotal': total, 'recordsFiltered': total_filtered, 'data': data})


# ---------------------------------------------------------------------------
# IP VIP TLS Check
# ---------------------------------------------------------------------------

@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def ip_vip_tls_check(request):
    """
    Node-to-VIP TLS lookup.

    Given one or more node IPs, finds:
      Pool members  → Pool → VIP (same ltm_fqdn, via default_pool)
      VIP profiles JSON (context = 'client-side' or 'all') → ClientSSLProfile lookup

    Returns a flat table with one row per (IP, VIP, profile):
      searched_ip, ltm_fqdn, vip_name, destination_address, destination_port,
      profile_name, profile_context, profile_tls (YES/NO)
    """
    from collections import defaultdict
    from ..models import ClientSSLProfile

    ips_raw  = ''
    results  = []
    searched = False

    if request.method == 'POST':
        ips_raw  = request.POST.get('ips', '')
        searched = True
        terms    = [ip.strip() for ip in ips_raw.splitlines() if ip.strip()]

        if terms:
            # 1. Find all pools containing any searched IP — single query via combined Q()
            terms_set = set(terms)
            q = Q()
            for ip in terms:
                q |= Q(members__contains=[{'address': ip}])
            ip_pools = {ip: [] for ip in terms}
            for pool in Pool.objects.filter(q):
                for m in (pool.members or []):
                    if isinstance(m, dict) and m.get('address') in terms_set:
                        ip_pools[m['address']].append(pool)

            all_pool_full_paths = {p.full_path for pools in ip_pools.values() for p in pools}

            if all_pool_full_paths:
                # 2. Fetch all relevant VIPs in one query
                all_vips = list(
                    VIP.objects.filter(default_pool__in=all_pool_full_paths)
                    .exclude(profiles__isnull=True).exclude(profiles='')
                )

                # 3. Pre-build TLS profile set: (ltm_fqdn, name) → exists
                relevant_ltm_fqdns = {v.ltm_fqdn for v in all_vips}
                tls_profile_set = set(
                    ClientSSLProfile.objects
                    .filter(ltm_fqdn__in=relevant_ltm_fqdns)
                    .values_list('ltm_fqdn', 'name')
                )

                # 4. Index VIPs by (default_pool, ltm_fqdn)
                vip_index = defaultdict(list)
                for vip in all_vips:
                    vip_index[(vip.default_pool, vip.ltm_fqdn)].append(vip)

                # 5. Build flat result rows
                for ip in terms:
                    pools = ip_pools.get(ip, [])
                    if not pools:
                        results.append({'searched_ip': ip, 'not_found': True})
                        continue

                    found_any = False
                    for pool in pools:
                        vips = vip_index.get((pool.full_path, pool.ltm_fqdn), [])
                        for vip in vips:
                            profiles_data = vip.profiles if isinstance(vip.profiles, list) else []
                            if not profiles_data:
                                continue

                            for p in profiles_data:
                                if not isinstance(p, dict):
                                    continue
                                context = p.get('context', '')
                                if context not in ('client-side', 'all'):
                                    continue
                                profile_name = p.get('name', '')
                                if not profile_name:
                                    continue

                                results.append({
                                    'searched_ip':         ip,
                                    'not_found':           False,
                                    'ltm_fqdn':            vip.ltm_fqdn or '-',
                                    'vip_name':            vip.name or '-',
                                    'destination_address': vip.destination_address or '-',
                                    'destination_port':    vip.destination_port or '-',
                                    'profile_name':        profile_name,
                                    'profile_context':     context,
                                    'profile_tls':         'YES' if (vip.ltm_fqdn, profile_name) in tls_profile_set else 'NO',
                                })
                                found_any = True

                    if not found_any:
                        results.append({'searched_ip': ip, 'not_found': True})
            else:
                for ip in terms:
                    results.append({'searched_ip': ip, 'not_found': True})

            # Post-process: compute cert_installed per (searched_ip, ltm_fqdn, vip_name)
            vip_has_tls = set()
            for r in results:
                if not r.get('not_found') and r.get('profile_tls') == 'YES':
                    vip_has_tls.add((r['searched_ip'], r['ltm_fqdn'], r['vip_name']))
            for r in results:
                if not r.get('not_found'):
                    key = (r['searched_ip'], r['ltm_fqdn'], r['vip_name'])
                    r['cert_installed'] = 'YES' if key in vip_has_tls else 'NO'

    return render(request, 'lb_manager/ip_vip_tls_check.html', {
        'ips_raw':  ips_raw,
        'results':  results,
        'searched': searched,
    })


# ---------------------------------------------------------------------------
# VIP Historical Snapshots
# ---------------------------------------------------------------------------

class LBVIPHistoricalListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """VIP historical snapshots list. Rows are loaded via lb_vip_historical_data."""

    model = LBVIPHistorical
    template_name = 'lb_manager/lb_vip_historical_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_lbviphistorical'
    raise_exception = True

    def get_queryset(self):
        return LBVIPHistorical.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'VIPs Historical'
        ctx['page_icon'] = 'fa-clock-rotate-left'
        ctx['total'] = LBVIPHistorical.objects.count()
        return ctx


@login_required
@permission_required('lb_manager.view_lbviphistorical', raise_exception=True)
def lb_vip_historical_data(request):
    """
    AJAX endpoint — returns paginated VIP Historical rows as JSON for DataTables.

    Requires at least one of: name, destination, ltm_fqdn.
    Returns an empty dataset when no filter is provided (deferred load pattern).
    """

    # Column index → model field name (col 0 = id, col 19 = date)
    COLUMNS = [
        'id', 'name', 'full_path', 'destination_address', 'destination_port',
        'destination', 'protocol', 'type', 'enabled', 'default_pool',
        'snat_type', 'snat_pool', 'source_address', 'connection_limit',
        'persistence_profile', 'profiles', 'policies', 'description',
        'ltm_fqdn', 'date',
    ]
    draw    = int(request.GET.get('draw', 1))
    start   = int(request.GET.get('start', 0))
    length  = _dt_length(request)
    col_idx = int(request.GET.get('order[0][column]', 19))
    col_dir = request.GET.get('order[0][dir]', 'desc')

    f_name   = request.GET.get('f_name', '').strip()
    f_dest   = request.GET.get('f_dest', '').strip()
    f_ltm    = request.GET.get('f_ltm', '').strip()

    # Return empty result when no filter is provided
    if not f_name and not f_dest and not f_ltm:
        return JsonResponse({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

    qs = LBVIPHistorical.objects.all()
    if f_name:
        qs = qs.filter(name=f_name)
    if f_dest:
        qs = qs.filter(destination_address=f_dest)
    if f_ltm:
        qs = qs.filter(ltm_fqdn=f_ltm)

    total    = qs.count()
    filtered = total
    order_col = COLUMNS[col_idx] if col_idx < len(COLUMNS) else 'date'
    if col_dir == 'desc':
        order_col = f'-{order_col}'
    qs = qs.order_by(order_col)[start:start + length]

    def _v(val):
        return val if val not in (None, '') else '-'

    data = [
        [
            h.id,
            _v(h.name),
            _v(h.full_path),
            _v(h.destination_address),
            _v(h.destination_port),
            _v(h.destination),
            _v(h.protocol),
            _v(h.type),
            _v(h.enabled),
            _v(h.default_pool),
            _v(h.snat_type),
            _v(h.snat_pool),
            _v(h.source_address),
            _v(h.connection_limit),
            _v(h.persistence_profile),
            _v(h.profiles),
            _v(h.policies),
            _v(h.description),
            _v(h.ltm_fqdn),
            str(h.date)[:10] if h.date else '-',
        ]
        for h in qs
    ]
    return JsonResponse({'draw': draw, 'recordsTotal': total, 'recordsFiltered': filtered, 'data': data})
