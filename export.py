"""
File download views — exporta datos de VIPs filtrados por empresa.

Soporta CSV, Excel y JSON con las mismas columnas que VIP Lookup
más LB Method (Pool.lb_method), que no está en el export estándar.
"""
import csv
import io
import json
from datetime import date

from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import render

from ..models import Company, LBGuest, LBPhysical, Pool, VIP

__all__ = ['vip_download']

# ── Constantes ────────────────────────────────────────────────────────────────

HEADERS = [
    'LTM FQDN',
    'VIP Name',
    'Destination IP',
    'Port',
    'Protocol',
    'Enabled',
    'Availability',
    'Default Pool',
    'LB Method',
    'Pool Members',
    'Monitors',
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _members_str(pool) -> str:
    """Builds a pipe-separated string of pool member names and addresses."""
    if not pool:
        return '-'
    raw = pool.members or []
    parts = []
    for m in raw:
        if isinstance(m, dict):
            fp   = m.get('full_path', '')
            port = fp.rsplit(':', 1)[-1] if ':' in fp else ''
            addr = m.get('address', '-')
            name = m.get('name', '-')
            parts.append(f'{name} ({addr}{":" + port if port else ""})')
    return ' | '.join(parts) if parts else '-'


def _monitors_str(pool) -> str:
    """Comma-separated monitor list from a Pool object."""
    if not pool:
        return '-'
    raw = pool.monitors or []
    return ', '.join(raw) if isinstance(raw, list) else '-'


def _build_rows(company_code: str) -> list[list]:
    """
    Queries VIPs (optionally filtered by company) and their pools.

    Args:
        company_code: client_code of the company, or '' for all.

    Returns:
        List of rows (one per VIP) matching HEADERS order.
    """
    if company_code:
        fqdns = set(
            LBGuest.objects.filter(
                company__client_code=company_code
            ).values_list('device', flat=True)
        ) | set(
            LBPhysical.objects.filter(
                company__client_code=company_code
            ).values_list('device', flat=True)
        )
        vips = list(VIP.objects.filter(ltm_fqdn__in=fqdns).order_by('ltm_fqdn', 'name'))
    else:
        vips = list(VIP.objects.order_by('ltm_fqdn', 'name'))

    pool_keys = {(v.default_pool, v.ltm_fqdn) for v in vips if v.default_pool}
    pool_map  = {
        (p.full_path, p.ltm_fqdn): p
        for p in Pool.objects.filter(
            full_path__in={k[0] for k in pool_keys},
            ltm_fqdn__in={k[1]  for k in pool_keys},
        )
    } if pool_keys else {}

    rows = []
    for vip in vips:
        pool = pool_map.get((vip.default_pool, vip.ltm_fqdn))
        rows.append([
            vip.ltm_fqdn              or '-',
            vip.name                  or '-',
            vip.destination_address   or '-',
            vip.destination_port      or '-',
            vip.protocol              or '-',
            vip.enabled               or '-',
            vip.availability_status   or '-',
            vip.default_pool          or '-',
            (pool.lb_method           or '-') if pool else '-',
            _members_str(pool),
            _monitors_str(pool),
        ])
    return rows


def _filename(company_code: str, ext: str) -> str:
    """Generates a safe filename like  vips_ACME_2026-05-16.xlsx."""
    slug = company_code.replace(' ', '_') if company_code else 'all'
    return f'vips_{slug}_{date.today().isoformat()}.{ext}'


# ── Generadores de respuesta ──────────────────────────────────────────────────

def _export_csv(rows: list, filename: str) -> HttpResponse:
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.write('﻿')  # BOM for Excel compatibility
    writer = csv.writer(resp)
    writer.writerow(HEADERS)
    writer.writerows(rows)
    return resp


def _export_excel(rows: list, filename: str) -> HttpResponse:
    from openpyxl import Workbook  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Alignment, Font, PatternFill  # pylint: disable=import-outside-toplevel

    wb = Workbook()
    ws = wb.active
    ws.title = 'VIPs'

    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, hdr in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _export_json(rows: list, filename: str) -> HttpResponse:
    data = [dict(zip(HEADERS, row)) for row in rows]
    resp = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ── Vista principal ───────────────────────────────────────────────────────────

@login_required
@permission_required('lb_manager.view_vip', raise_exception=True)
def vip_download(request):
    """
    Descarga el inventario de VIPs filtrado por empresa.

    GET  → página con selector de empresa y formato.
    POST → genera y descarga el archivo en el formato elegido.

    Args:
        company: client_code de la empresa, o '' para todas.
        fmt:     'csv' | 'excel' | 'json'.
    """
    companies = Company.objects.order_by('name')

    if request.method == 'POST':
        company_code = request.POST.get('company', '').strip()
        fmt          = request.POST.get('fmt', 'excel')

        rows     = _build_rows(company_code)
        filename = _filename(company_code, {'csv': 'csv', 'excel': 'xlsx', 'json': 'json'}.get(fmt, 'xlsx'))

        if fmt == 'csv':
            return _export_csv(rows, filename)
        if fmt == 'json':
            return _export_json(rows, filename)
        return _export_excel(rows, filename)

    return render(request, 'lb_manager/downloads.html', {
        'companies': companies,
        'headers':   HEADERS,
    })
