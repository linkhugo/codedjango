"""
SSL Certificate and Profile views.
"""

import io
import socket
from datetime import datetime, timezone as dt_timezone, timedelta

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.generic import ListView

from ..models import SSLCert, ClientSSLProfile, VIP, Pool
from .utils import _dt_params, _parse_cert_expiration, apply_search_filter


class SSLCertListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Renders the SSL Certificate list page shell. Rows are loaded via ssl_cert_data."""

    model = SSLCert
    template_name = 'lb_manager/ssl_cert_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_sslcert'
    raise_exception = True

    def get_queryset(self):
        return SSLCert.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'SSL Certificates'
        ctx['page_icon'] = 'fa-certificate'
        ctx['total'] = SSLCert.objects.count()
        return ctx


@login_required
@permission_required('lb_manager.view_sslcert', raise_exception=True)
def ssl_cert_data(request):
    """AJAX endpoint — returns paginated SSL Certificate rows as JSON for DataTables."""

    COLUMNS = ['id', 'name', 'expiration_date', 'subject', 'issuer', 'key_type', 'key_size',
               'fingerprint', 'sha1_checksum', 'is_bundle', 'full_path', 'ltm_fqdn',
               'create_time', 'last_update_time', 'expiration_timestamp', 'subject_alternative_name']
    draw, start, length, search, col_idx, col_dir = _dt_params(request)

    qs = SSLCert.objects.all()
    total = qs.count()
    qs = apply_search_filter(qs, search, [
        'name', 'subject', 'issuer', 'ltm_fqdn', 'full_path', 'fingerprint',
    ])
    filtered = qs.count()
    order_col = COLUMNS[col_idx] if col_idx < len(COLUMNS) else 'id'
    if col_dir == 'desc':
        order_col = f'-{order_col}'
    qs = qs.order_by(order_col)[start:start + length]
    data = [[
        c.id,
        c.name or '-',
        c.expiration_date or '-',
        (c.subject or '-')[:100],
        (c.issuer or '-')[:100],
        c.key_type or '-',
        c.key_size if c.key_size is not None else '-',
        c.fingerprint or '-',
        c.sha1_checksum or '-',
        c.is_bundle or '-',
        c.full_path or '-',
        c.ltm_fqdn or '-',
        c.create_time or '-',
        c.last_update_time or '-',
        c.expiration_timestamp if c.expiration_timestamp is not None else '-',
        c.subject_alternative_name or '-',
    ] for c in qs]
    return JsonResponse({'draw': draw, 'recordsTotal': total, 'recordsFiltered': filtered, 'data': data})


class SSLProfileListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Renders the Client SSL Profile list page shell. Rows are loaded via ssl_profile_data."""

    model = ClientSSLProfile
    template_name = 'lb_manager/ssl_profile_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_clientsslprofile'
    raise_exception = True

    def get_queryset(self):
        return ClientSSLProfile.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Client SSL Profiles'
        ctx['page_icon'] = 'fa-shield-halved'
        ctx['total'] = ClientSSLProfile.objects.count()
        return ctx


@login_required
@permission_required('lb_manager.view_clientsslprofile', raise_exception=True)
def ssl_profile_data(request):
    """AJAX endpoint — returns paginated Client SSL Profile rows as JSON for DataTables."""

    COLUMNS = ['id', 'name', 'certificate_file', 'chain_file',
               'allow_non_ssl', 'parent', 'ltm_fqdn']
    draw, start, length, search, col_idx, col_dir = _dt_params(request)

    qs = ClientSSLProfile.objects.all()
    total = qs.count()
    qs = apply_search_filter(qs, search, ['name', 'certificate_file', 'ltm_fqdn', 'parent'])
    filtered = qs.count()
    order_col = COLUMNS[col_idx] if col_idx < len(COLUMNS) else 'id'
    if col_dir == 'desc':
        order_col = f'-{order_col}'
    qs = qs.order_by(order_col)[start:start + length]
    data = [[p.id, p.name or '-', p.certificate_file or '-', p.chain_file or '-',
             p.allow_non_ssl or '-', p.parent or '-',
             p.ltm_fqdn or '-'] for p in qs]
    return JsonResponse({'draw': draw, 'recordsTotal': total, 'recordsFiltered': filtered, 'data': data})


@login_required
@permission_required('lb_manager.view_sslcert', raise_exception=True)
def vip_expired_ssl(request):
    """
    Shows VIPs using SSL Certificates that are expired (red) or expiring within 30 days (yellow).

    Chain:
      1. vips.profiles (JSON) → entry where context="clientside" → name
      2. client_ssl_profiles.name = step1.name  (same ltm_fqdn)
      3. ssl_certs.full_path = client_ssl_profiles.certificate_file  (same ltm_fqdn)
      4. Check ssl_certs.expiration_date / expiration_timestamp
    """
    today       = datetime.now(dt_timezone.utc).date()
    soon_date   = today + timedelta(days=30)

    # ── 1. Certs that are expired OR expiring within 30 days ─────────────────
    relevant_ids = []
    for cert in SSLCert.objects.exclude(
            expiration_timestamp__isnull=True, expiration_date__isnull=True):
        exp = _parse_cert_expiration(cert)
        if exp is not None and exp <= soon_date:
            relevant_ids.append(cert.id)

    if not relevant_ids:
        return render(request, 'lb_manager/ssl_expired_vips.html', {'results': [], 'total': 0})

    relevant_certs = SSLCert.objects.filter(id__in=relevant_ids)

    # ── 2. Build lookup: (ltm_fqdn, full_path) → cert ────────────────────────
    cert_map = {}
    for cert in relevant_certs:
        if cert.full_path:
            cert_map[(cert.ltm_fqdn, cert.full_path)] = cert

    relevant_full_paths = {fp for (_, fp) in cert_map}

    # ── 3. SSL profiles: certificate_file = full_path, same ltm_fqdn ─────────
    #    Build: (ltm_fqdn, profile_name) → (profile, cert)
    profile_cert_map = {}
    for profile in ClientSSLProfile.objects.filter(
            certificate_file__in=relevant_full_paths
    ).exclude(name__isnull=True).exclude(name=''):
        cert = cert_map.get((profile.ltm_fqdn, profile.certificate_file))
        if cert:
            profile_cert_map[(profile.ltm_fqdn, profile.name)] = (profile, cert)

    if not profile_cert_map:
        return render(request, 'lb_manager/ssl_expired_vips.html', {'results': [], 'total': 0})

    # ── 4. VIPs: parse profiles JSON, find context="clientside", match name ──
    results = []
    seen    = set()

    for vip in VIP.objects.exclude(profiles__isnull=True).exclude(profiles=''):
        profiles_data = vip.profiles if isinstance(vip.profiles, list) else []

        for p in profiles_data:
            if not isinstance(p, dict):
                continue
            if p.get('context') != 'client-side':
                continue
            pname = p.get('name', '')
            if not pname:
                continue

            entry = profile_cert_map.get((vip.ltm_fqdn, pname))
            if not entry:
                continue

            key = (vip.id, pname)
            if key in seen:
                continue
            seen.add(key)

            profile, cert = entry
            exp = _parse_cert_expiration(cert)
            if exp is None:
                continue

            days_diff = (today - exp).days
            status    = 'expired' if exp < today else 'expiring_soon'

            results.append({
                'vip_name':             vip.name or '-',
                'vip_destination':      vip.destination_address or '-',
                'vip_destination_port': vip.destination_port,
                'vip_enabled':          vip.enabled or '-',
                'ltm_fqdn':             vip.ltm_fqdn or '-',
                'ssl_profile_name':     pname,
                'cert_name':            cert.full_path or cert.name or profile.certificate_file,
                'cert_subject':         cert.subject or '-',
                'cert_expiration':      cert.expiration_date or '-',
                'days_diff':            days_diff,
                'days_abs':             abs(days_diff),
                'status':               status,
            })

    # Sort: expired first (largest days_diff), then expiring_soon by days remaining
    results.sort(key=lambda r: (-r['days_diff']))

    expired_count = sum(1 for r in results if r['status'] == 'expired')
    soon_count    = sum(1 for r in results if r['status'] == 'expiring_soon')

    return render(request, 'lb_manager/ssl_expired_vips.html', {
        'results':       results,
        'total':         len(results),
        'expired_count': expired_count,
        'soon_count':    soon_count,
    })


# ── SSL Hostname Search ───────────────────────────────────────────────────────

@login_required
@permission_required('lb_manager.view_sslcert', raise_exception=True)
def ssl_hostname_search(request):
    """
    Search SSL certs and VIPs by hostname.

    For each hostname the lookup chain is:
      SSLCert (subject / SAN) → ClientSSLProfile → VIP (profiles JSON) → Pool

    Fallback when no cert match: resolve hostname to IP → VIP.destination_address

    Supports Excel export via POST param export=1.
    """
    results = []
    searched = False
    raw_input = ''

    if request.method == 'POST':
        raw_input = request.POST.get('hostnames', '')
        do_export = request.POST.get('export') == '1'
        hostname_list = [h.strip() for h in raw_input.splitlines() if h.strip()]
        searched = bool(hostname_list)

        for hostname in hostname_list:
            results.extend(_search_hostname(hostname))

        if do_export and results:
            return _hostname_excel_response(results)

    return render(request, 'lb_manager/ssl_hostname_search.html', {
        'results':    results,
        'searched':   searched,
        'raw_input':  raw_input,
        'total':      len(results),
    })


def _search_hostname(hostname):
    """Return a list of result-row dicts for one hostname."""
    rows = []

    # ── 1. Find matching SSL certs ─────────────────────────────────────────
    certs = list(SSLCert.objects.filter(
        Q(subject__icontains=f'CN={hostname}') |
        Q(subject_alternative_name__icontains=f'DNS:{hostname}')
    ))

    if certs:
        # (ltm_fqdn, full_path) → cert
        cert_map = {(c.ltm_fqdn, c.full_path): c for c in certs if c.full_path}

        # ── 2. Find SSL profiles using those certs ─────────────────────────
        profiles = list(ClientSSLProfile.objects.filter(
            certificate_file__in={fp for (_, fp) in cert_map}
        ).exclude(name__isnull=True).exclude(name=''))

        # (ltm_fqdn, profile_name) → (profile, cert)
        profile_cert = {}
        for p in profiles:
            cert = cert_map.get((p.ltm_fqdn, p.certificate_file))
            if cert:
                profile_cert[(p.ltm_fqdn, p.name)] = (p, cert)

        if profile_cert:
            # ── 3. Find VIPs containing those profile names ────────────────
            q = Q()
            for name in {pname for (_, pname) in profile_cert}:
                q |= Q(profiles__icontains=name)
            vips = list(VIP.objects.filter(q).exclude(profiles__isnull=True))

            # pre-fetch pools
            pool_keys = {(v.ltm_fqdn, v.default_pool) for v in vips if v.default_pool}
            pool_map = {}
            if pool_keys:
                ltms = {k[0] for k in pool_keys}
                paths = {k[1] for k in pool_keys}
                for pool in Pool.objects.filter(ltm_fqdn__in=ltms, full_path__in=paths):
                    pool_map[(pool.ltm_fqdn, pool.full_path)] = pool

            seen = set()
            for vip in vips:
                pdata = vip.profiles if isinstance(vip.profiles, list) else []
                for p in pdata:
                    if not isinstance(p, dict):
                        continue
                    if p.get('context') not in ('client-side', 'clientside'):
                        continue
                    pname = p.get('name', '')
                    if not pname:
                        continue
                    entry = profile_cert.get((vip.ltm_fqdn, pname))
                    if not entry:
                        continue
                    key = (vip.id, pname)
                    if key in seen:
                        continue
                    seen.add(key)
                    _, cert = entry
                    pool = pool_map.get((vip.ltm_fqdn, vip.default_pool))
                    m_names, m_addrs = _fmt_members(pool)
                    rows.append({
                        'dns':                 hostname,
                        'nombre_granja':       vip.name or '-',
                        'balanceador':         vip.ltm_fqdn or '-',
                        'ip_granja':           vip.destination_address or '-',
                        'puerto_granja':       vip.destination_port or '-',
                        'common_name':         cert.subject or '-',
                        'alter_names':         cert.subject_alternative_name or '-',
                        'fecha_expiracion':    cert.expiration_date or '-',
                        'profiles':            _fmt_profiles(pdata),
                        'pool_members':        m_names,
                        'pool_members_address': m_addrs,
                        'tipo':                'SI',
                        'row_class':           'table-success',
                    })

    # ── Fallback: no cert found — try resolving IP and searching VIPs ─────────
    if not rows:
        resolved_ip = None
        try:
            resolved_ip = socket.gethostbyname(hostname)
        except (socket.gaierror, OSError):
            pass

        if resolved_ip:
            for vip in VIP.objects.filter(destination_address=resolved_ip):
                pool = Pool.objects.filter(
                    full_path=vip.default_pool, ltm_fqdn=vip.ltm_fqdn
                ).first()
                m_names, m_addrs = _fmt_members(pool)
                pdata = vip.profiles if isinstance(vip.profiles, list) else []
                rows.append({
                    'dns':                 hostname,
                    'nombre_granja':       vip.name or '-',
                    'balanceador':         vip.ltm_fqdn or '-',
                    'ip_granja':           vip.destination_address or '-',
                    'puerto_granja':       vip.destination_port or '-',
                    'common_name':         '-',
                    'alter_names':         '-',
                    'fecha_expiracion':    '-',
                    'profiles':            _fmt_profiles(pdata),
                    'pool_members':        m_names,
                    'pool_members_address': m_addrs,
                    'tipo':                'NO',
                    'row_class':           'table-warning',
                })

        if not rows:
            tipo = (
                'Granja Externa — validar manualmente'
                if '.corp' not in hostname.lower()
                else 'Sin granja registrada'
            )
            rows.append({
                'dns':                 hostname,
                'nombre_granja':       'No se encontró información',
                'balanceador':         '-',
                'ip_granja':           str(resolved_ip) if resolved_ip else '-',
                'puerto_granja':       '-',
                'common_name':         '-',
                'alter_names':         '-',
                'fecha_expiracion':    '-',
                'profiles':            '-',
                'pool_members':        '-',
                'pool_members_address': '-',
                'tipo':                tipo,
                'row_class':           'table-danger',
            })

    return rows


def _fmt_members(pool):
    """Return (names_str, addresses_str) from a Pool object."""
    if not pool or not pool.members:
        return '-', '-'
    names = [m.get('name', '') for m in pool.members if isinstance(m, dict)]
    addrs = [m.get('address', '') for m in pool.members if isinstance(m, dict)]
    return (', '.join(filter(None, names)) or '-',
            ', '.join(filter(None, addrs)) or '-')


def _fmt_profiles(pdata):
    """Return comma-separated profile names from a parsed profiles list."""
    if not isinstance(pdata, list):
        return '-'
    names = [p.get('name', '') for p in pdata if isinstance(p, dict)]
    return ', '.join(filter(None, names)) or '-'


def _hostname_excel_response(results):
    """Build and return an Excel file download for hostname search results."""
    from openpyxl import Workbook  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Font, PatternFill, Alignment  # pylint: disable=import-outside-toplevel

    headers = [
        'DNS', 'Nombre_Granja', 'Balanceador', 'IP_Granja', 'Puerto_Granja',
        'Common_Name', 'AlterNames', 'Fecha_Expiracion',
        'Profiles', 'Pool_Members', 'Pool_Members_Address', 'Tipo',
    ]
    keys = [
        'dns', 'nombre_granja', 'balanceador', 'ip_granja', 'puerto_granja',
        'common_name', 'alter_names', 'fecha_expiracion',
        'profiles', 'pool_members', 'pool_members_address', 'tipo',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hostname SSL Search'

    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    _NO_POOL_MSG = 'Esta VIP puede tener una iRule o Policy, contactar al Administrador para su revision'

    for row_idx, row in enumerate(results, 2):
        for col, key in enumerate(keys, 1):
            if key == 'pool_members' and row.get('pool_members') == '-' and row.get('row_class') != 'table-danger':
                ws.cell(row=row_idx, column=col, value=_NO_POOL_MSG)
            else:
                ws.cell(row=row_idx, column=col, value=str(row.get(key, '-')))

    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="hostname_ssl_search.xlsx"'
    return resp


# ── SSL Dashboard ─────────────────────────────────────────────────────────────

@login_required
@permission_required('lb_manager.view_sslcert', raise_exception=True)
def ssl_dashboard(request):
    """
    Consolidated SSL dashboard.

    Only includes SSLCerts that are actively referenced by a ClientSSLProfile
    which is in turn attached to at least one VIP (client-side context).

    Buckets each qualifying cert into one of four tiers based on days to expiration:
      expired   — already past due
      critical  — expires within 30 days
      warning   — expires within 31–90 days
      ok        — more than 90 days remaining (or no expiration data)

    Builds a 12-month timeline of upcoming expirations for a bar chart and
    a filterable table of all non-OK certs.
    """
    import calendar as _cal

    today = datetime.now(dt_timezone.utc).date()

    # ── Thresholds ────────────────────────────────────────────────────────────
    CRITICAL_DAYS = 30
    WARNING_DAYS  = 90

    # ── Step 1: collect (ltm_fqdn, profile_name) from VIPs (client-side only) ─
    vip_profile_keys: set[tuple[str, str]] = set()
    for vip in VIP.objects.only('ltm_fqdn', 'profiles'):
        profiles_list = vip.profiles if isinstance(vip.profiles, list) else []
        for p in profiles_list:
            if isinstance(p, dict) and p.get('context') == 'client-side':
                name = p.get('name', '').strip()
                if name:
                    vip_profile_keys.add((vip.ltm_fqdn or '', name))

    # ── Step 2: find (ltm_fqdn, certificate_file) for those profiles ─────────
    cert_keys: set[tuple[str, str]] = set()
    for profile in ClientSSLProfile.objects.only('name', 'ltm_fqdn', 'certificate_file'):
        if (profile.ltm_fqdn or '', profile.name or '') in vip_profile_keys:
            cert_file = (profile.certificate_file or '').strip()
            if cert_file:
                cert_keys.add((profile.ltm_fqdn or '', cert_file))

    # ── Bucket counters and row list ──────────────────────────────────────────
    counts = {'expired': 0, 'critical': 0, 'warning': 0, 'ok': 0}
    rows: list[dict] = []
    total = 0

    # ── 12-month timeline: month_label → count of certs expiring that month ──
    timeline: dict[str, int] = {}
    for i in range(12):
        m = today.month + i
        y = today.year + (m - 1) // 12
        m = (m - 1) % 12 + 1
        timeline[f'{y}-{m:02d}'] = 0

    # ── Step 3: iterate only certs linked to a VIP via a ClientSSLProfile ────
    for cert in SSLCert.objects.exclude(
        expiration_timestamp__isnull=True, expiration_date__isnull=True
    ).only(
        'id', 'name', 'full_path', 'ltm_fqdn',
        'expiration_date', 'expiration_timestamp',
        'issuer', 'subject', 'key_type', 'key_size',
    ):
        if (cert.ltm_fqdn or '', cert.full_path or '') not in cert_keys:
            continue

        total += 1
        exp = _parse_cert_expiration(cert)
        if exp is None:
            counts['ok'] += 1
            continue

        days_left = (exp - today).days

        if days_left < 0:
            tier = 'expired'
        elif days_left <= CRITICAL_DAYS:
            tier = 'critical'
        elif days_left <= WARNING_DAYS:
            tier = 'warning'
        else:
            tier = 'ok'

        counts[tier] += 1

        # Add to timeline if expires within the next 12 months
        if 0 <= days_left <= 365:
            key = f'{exp.year}-{exp.month:02d}'
            if key in timeline:
                timeline[key] += 1

        if tier != 'ok':
            rows.append({
                'cert_name':   cert.name or cert.full_path or '-',
                'full_path':   cert.full_path or '-',
                'ltm_fqdn':    cert.ltm_fqdn or '-',
                'issuer':      (cert.issuer or '-')[:80],
                'subject':     (cert.subject or '-')[:80],
                'key_type':    cert.key_type or '-',
                'key_size':    cert.key_size or '-',
                'exp_date':    exp.strftime('%Y-%m-%d'),
                'days_left':   days_left,
                'tier':        tier,
            })

    # Sort: expired first (most overdue), then by days_left ascending
    rows.sort(key=lambda r: r['days_left'])

    # Build timeline labels and data for Chart.js
    timeline_labels = []
    timeline_data   = []
    for key, cnt in timeline.items():
        y, m = key.split('-')
        timeline_labels.append(_cal.month_abbr[int(m)] + ' ' + y)
        timeline_data.append(cnt)

    return render(request, 'lb_manager/ssl_dashboard.html', {
        'counts':           counts,
        'total':            total,
        'rows':             rows,
        'timeline_labels':  timeline_labels,
        'timeline_data':    timeline_data,
        'today':            today.strftime('%Y-%m-%d'),
    })
