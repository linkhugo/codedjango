"""
Infrastructure views: LBPhysical, LBGuest, Company, Datacenter, inventory, and CMDB validation.
"""

import csv
import io
import re

from django.contrib import messages
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.contenttypes.models import ContentType
from django.db import connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils import timezone as dj_timezone
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from ..models import (
    Company, Datacenter, LBPhysical, LBGuest, LBDeviceChangeLog,
    LBDecommissioned, SiteSettings, CMDBFieldConfig, AnsibleInventoryFile,
)
from ..forms import CompanyForm, DatacenterForm, LBPhysicalForm, LBGuestForm
from .utils import (
    _dt_length, _log_wiki_change, _record_device_changes,
    _IP_RESET_WARNING_MSG, _CMDB_COLS, _RESULT_COLS, _OK, _NA, _cmdb_compare_row,
)


class LBPhysicalListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Physical LB inventory list. Supports optional ?environment= filter in the URL
    to pre-filter the table to a specific environment (DRP, PRE-PRODUCTION, PRODUCTION).
    """

    model = LBPhysical
    template_name = 'lb_manager/lb_physical_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_lbphysical'
    raise_exception = True

    def get_queryset(self):
        qs = super().get_queryset().select_related('company', 'datacenter')
        env = self.request.GET.get('environment')
        if env:
            qs = qs.filter(environment__iexact=env)
        company = self.request.GET.get('company')
        if company:
            qs = qs.filter(company__client_code__iexact=company)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Physical Load Balancers'
        ctx['page_icon'] = 'fa-hdd'
        ctx['environment'] = self.request.GET.get('environment', '')
        ctx['company_filter'] = self.request.GET.get('company', '')
        ctx['companies'] = Company.objects.order_by('client_code').values_list('client_code', flat=True)
        return ctx


class LBGuestListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Virtual (guest) LB inventory list."""

    model = LBGuest
    template_name = 'lb_manager/lb_guest_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_lbguest'
    raise_exception = True

    def get_queryset(self):
        qs = super().get_queryset().select_related('company')
        company = self.request.GET.get('company')
        if company:
            qs = qs.filter(company__client_code__iexact=company)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Guest Load Balancers'
        ctx['page_icon'] = 'fa-cloud-upload-alt'
        ctx['company_filter'] = self.request.GET.get('company', '')
        ctx['companies'] = Company.objects.order_by('client_code').values_list('client_code', flat=True)
        return ctx


class CompanyListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Client/company catalog list."""

    model = Company
    template_name = 'lb_manager/company_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_company'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Companies'
        ctx['page_icon'] = 'fa-building'
        return ctx


class DatacenterListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Datacenter location catalog list."""

    model = Datacenter
    template_name = 'lb_manager/datacenter_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_datacenter'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Datacenters'
        ctx['page_icon'] = 'fa-warehouse'
        return ctx


# ============================================================
#  COMPANY CRUD
#  Requires permission: lb_manager.add/change/delete_company
#  Returns HTTP 403 (not redirect) if the user lacks the permission.
# ============================================================
class CompanyCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new Company record. Requires the 'add_company' permission."""
    model = Company
    form_class = CompanyForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('company_list')
    permission_required = 'lb_manager.add_company'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add Company'
        ctx['list_url'] = reverse_lazy('company_list')
        ctx['list_label'] = 'Companies'
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Company created successfully.')
        return super().form_valid(form)


class CompanyUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Edit an existing Company record. Requires the 'change_company' permission."""

    model = Company
    form_class = CompanyForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('company_list')
    permission_required = 'lb_manager.change_company'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Company: {self.object}'
        ctx['list_url'] = reverse_lazy('company_list')
        ctx['list_label'] = 'Companies'
        ctx['is_edit'] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Company updated successfully.')
        return super().form_valid(form)


class CompanyDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a Company record after confirmation. Requires the 'delete_company' permission."""

    model = Company
    template_name = 'lb_manager/crud_confirm_delete.html'
    success_url = reverse_lazy('company_list')
    permission_required = 'lb_manager.delete_company'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Delete Company'
        ctx['list_url'] = reverse_lazy('company_list')
        ctx['list_label'] = 'Companies'
        ctx['object_name'] = str(self.object)
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Company deleted successfully.')
        return super().form_valid(form)


# ============================================================
#  DATACENTER CRUD
# ============================================================
class DatacenterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new Datacenter record. Requires the 'add_datacenter' permission."""

    model = Datacenter
    form_class = DatacenterForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('datacenter_list')
    permission_required = 'lb_manager.add_datacenter'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add Datacenter'
        ctx['list_url'] = reverse_lazy('datacenter_list')
        ctx['list_label'] = 'Datacenters'
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Datacenter created successfully.')
        return super().form_valid(form)


class DatacenterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Edit an existing Datacenter record. Requires the 'change_datacenter' permission."""

    model = Datacenter
    form_class = DatacenterForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('datacenter_list')
    permission_required = 'lb_manager.change_datacenter'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Datacenter: {self.object}'
        ctx['list_url'] = reverse_lazy('datacenter_list')
        ctx['list_label'] = 'Datacenters'
        ctx['is_edit'] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Datacenter updated successfully.')
        return super().form_valid(form)


class DatacenterDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a Datacenter record after confirmation. Requires the 'delete_datacenter' permission."""

    model = Datacenter
    template_name = 'lb_manager/crud_confirm_delete.html'
    success_url = reverse_lazy('datacenter_list')
    permission_required = 'lb_manager.delete_datacenter'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Delete Datacenter'
        ctx['list_url'] = reverse_lazy('datacenter_list')
        ctx['list_label'] = 'Datacenters'
        ctx['object_name'] = str(self.object)
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Datacenter deleted successfully.')
        return super().form_valid(form)


# ============================================================
#  LB PHYSICAL CRUD
# ============================================================
class LBPhysicalCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Register a new physical F5 appliance. Requires the 'add_lbphysical' permission."""

    model = LBPhysical
    form_class = LBPhysicalForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('lb_physical_list')
    permission_required = 'lb_manager.add_lbphysical'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add Physical LB'
        ctx['list_url'] = reverse_lazy('lb_physical_list')
        ctx['list_label'] = 'Physical LBs'
        ctx['dado_de_alta_fields'] = ['monitoreo', 'cmdb_snmp', 'cyberark', 'user_cyberark', 'url_cyberark']
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Physical LB created successfully.')
        return super().form_valid(form)


class LBPhysicalUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Edit an existing physical F5 appliance record. Requires the 'change_lbphysical' permission."""

    model = LBPhysical
    form_class = LBPhysicalForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('lb_physical_list')
    permission_required = 'lb_manager.change_lbphysical'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Physical LB: {self.object.device}'
        ctx['list_url'] = reverse_lazy('lb_physical_list')
        ctx['list_label'] = 'Physical LBs'
        ctx['is_edit'] = True
        ctx['dado_de_alta_fields'] = ['monitoreo', 'cmdb_snmp', 'cyberark', 'user_cyberark', 'url_cyberark']
        return ctx

    def form_valid(self, form):
        try:
            old = LBPhysical.objects.select_related('company', 'datacenter').get(pk=self.object.pk)
        except LBPhysical.DoesNotExist:
            old = None
        response = super().form_valid(form)
        if old:
            _record_device_changes(self.request.user, old, self.object, 'physical', form.changed_data)
        _log_wiki_change(self.request.user, self.object, CHANGE)
        if getattr(form, '_ip_reset_warning', False):
            messages.warning(self.request, _IP_RESET_WARNING_MSG)
        messages.success(self.request, 'Physical LB updated successfully.')
        return response


class LBPhysicalDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a physical F5 appliance record after confirmation. Requires the 'delete_lbphysical' permission."""

    model = LBPhysical
    template_name = 'lb_manager/lb_decommission_confirm.html'
    success_url = reverse_lazy('lb_physical_list')
    permission_required = 'lb_manager.delete_lbphysical'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Decomisionar Physical LB'
        ctx['list_url'] = reverse_lazy('lb_physical_list')
        ctx['list_label'] = 'Physical LBs'
        ctx['object_name'] = self.object.device
        ctx['device_type'] = 'Physical'
        return ctx

    def form_valid(self, form):
        obj = self.get_object()
        LBDecommissioned.objects.create(
            device=obj.device,
            device_type='Physical',
            ci_id=obj.ci_id,
            mgmt_ip=obj.mgmt_ip,
            version=obj.version,
            model=obj.model,
            serial=obj.serial,
            environment=obj.environment,
            company=obj.company.client_code if obj.company else None,
            comments=self.request.POST.get('comments', '').strip() or None,
            decommissioned_by=self.request.user,
        )
        messages.success(self.request, f'Physical LB "{obj.device}" decomisado y eliminado.')
        return super().form_valid(form)


# ============================================================
#  LB GUEST CRUD
# ============================================================
class LBGuestCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Register a new virtual (guest) F5 instance. Requires the 'add_lbguest' permission."""

    model = LBGuest
    form_class = LBGuestForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('lb_guest_list')
    permission_required = 'lb_manager.add_lbguest'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add Guest LB'
        ctx['list_url'] = reverse_lazy('lb_guest_list')
        ctx['list_label'] = 'Guest LBs'
        ctx['dado_de_alta_fields'] = ['monitoreo', 'cmdb_snmp', 'cyberark', 'user_cyberark', 'url_cyberark']
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Guest LB created successfully.')
        return super().form_valid(form)


class LBGuestUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Edit an existing virtual F5 instance record. Requires the 'change_lbguest' permission."""

    model = LBGuest
    form_class = LBGuestForm
    template_name = 'lb_manager/crud_form.html'
    success_url = reverse_lazy('lb_guest_list')
    permission_required = 'lb_manager.change_lbguest'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Guest LB: {self.object.device}'
        ctx['list_url'] = reverse_lazy('lb_guest_list')
        ctx['list_label'] = 'Guest LBs'
        ctx['is_edit'] = True
        ctx['dado_de_alta_fields'] = ['monitoreo', 'cmdb_snmp', 'cyberark', 'user_cyberark', 'url_cyberark']
        return ctx

    def form_valid(self, form):
        try:
            old = LBGuest.objects.select_related('company').get(pk=self.object.pk)
        except LBGuest.DoesNotExist:
            old = None
        response = super().form_valid(form)
        if old:
            _record_device_changes(self.request.user, old, self.object, 'guest', form.changed_data)
        _log_wiki_change(self.request.user, self.object, CHANGE)
        if getattr(form, '_ip_reset_warning', False):
            messages.warning(self.request, _IP_RESET_WARNING_MSG)
        messages.success(self.request, 'Guest LB updated successfully.')
        return response


class LBGuestDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a virtual F5 instance record after confirmation. Requires the 'delete_lbguest' permission."""

    model = LBGuest
    template_name = 'lb_manager/lb_decommission_confirm.html'
    success_url = reverse_lazy('lb_guest_list')
    permission_required = 'lb_manager.delete_lbguest'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Decomisionar Guest LB'
        ctx['list_url'] = reverse_lazy('lb_guest_list')
        ctx['list_label'] = 'Guest LBs'
        ctx['object_name'] = self.object.device
        ctx['device_type'] = 'Guest'
        return ctx

    def form_valid(self, form):
        obj = self.get_object()
        LBDecommissioned.objects.create(
            device=obj.device,
            device_type='Guest',
            ci_id=obj.ci_id,
            mgmt_ip=obj.mgmt_ip,
            version=obj.version,
            model=obj.model,
            serial=obj.serial,
            environment=obj.environment,
            company=obj.company.client_code if obj.company else None,
            comments=self.request.POST.get('comments', '').strip() or None,
            decommissioned_by=self.request.user,
        )
        messages.success(self.request, f'Guest LB "{obj.device}" decomisado y eliminado.')
        return super().form_valid(form)


class LBDecommissionedListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Read-only log of all decommissioned LBPhysical and LBGuest devices."""

    model = LBDecommissioned
    template_name = 'lb_manager/lb_decommissioned_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_lbdecommissioned'
    raise_exception = True

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Equipos Decomisados'
        ctx['page_icon'] = 'fa-trash-can-arrow-up'
        return ctx


@login_required
@permission_required('lb_manager.view_lbdevicechangelog', raise_exception=True)
def lb_device_history(_request, device):
    """
    AJAX — return the field-level change history for a single LB device.

    Returns JSON ``{data: [...]}`` where each item has:
      timestamp, user, field, old_value, new_value
    """
    _lb_hist_cfg = SiteSettings.objects.first()
    _lb_hist_limit = _lb_hist_cfg.lb_device_history_limit if _lb_hist_cfg else 200
    entries = (
        LBDeviceChangeLog.objects
        .filter(device=device)
        .select_related('user')
        .order_by('-timestamp')[:_lb_hist_limit]
    )
    data = [
        {
            'timestamp': e.timestamp.strftime('%d/%m/%Y %H:%M'),
            'user': e.user.get_full_name() or e.user.username if e.user else '—',
            'field': e.field_name,
            'old': e.old_value or '—',
            'new': e.new_value or '—',
        }
        for e in entries
    ]
    return JsonResponse({'data': data})


# ============================================================
#  UNIFIED LB INVENTORY  (LBGuest UNION LBPhysical)
# ============================================================
class LBInventoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Unified inventory combining every LBGuest and every LBPhysical in one table.

    Replicates the SQL UNION query:
      - Part 1: All LBGuest rows, left-joined to LBPhysical by serial number
        so each guest inherits the physical host's datacenter, location,
        service date, vendor support, and virtualizador name.
      - Part 2: All LBPhysical rows (standalone appliances).

    Rows are loaded asynchronously via lb_inventory_data.
    """

    model = LBGuest  # dummy model — real data comes from the AJAX endpoint
    template_name = 'lb_manager/lb_inventory_list.html'
    context_object_name = 'objects'
    paginate_by = None
    permission_required = 'lb_manager.view_lbphysical'
    raise_exception = True

    def get_queryset(self):
        return LBGuest.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'LB Inventory'
        ctx['page_icon'] = 'fa-table-list'
        ctx['total'] = LBGuest.objects.count() + LBPhysical.objects.count()
        ctx['companies'] = Company.objects.order_by('client_code').values_list('client_code', flat=True)

        lb_ct_ids = list(ContentType.objects.filter(
            app_label='lb_manager', model__in=['lbguest', 'lbphysical']
        ).values_list('id', flat=True))
        _inv_cfg = SiteSettings.objects.first()
        _inv_limit = _inv_cfg.inventory_recent_changes_limit if _inv_cfg else 50
        recent = (
            LogEntry.objects
            .filter(content_type_id__in=lb_ct_ids)
            .select_related('user', 'content_type')
            .order_by('-action_time')[:_inv_limit]
        )
        ACTION_LABEL = {ADDITION: 'Agregado', CHANGE: 'Modificado', DELETION: 'Eliminado'}
        ACTION_COLOR = {ADDITION: 'success', CHANGE: 'primary', DELETION: 'danger'}
        ctx['inventory_changes'] = [
            {
                'timestamp': e.action_time,
                'user': e.user.get_full_name() or e.user.username if e.user else '—',
                'model': 'Physical LB' if e.content_type.model == 'lbphysical' else 'Guest LB',
                'device': e.object_repr,
                'action': ACTION_LABEL.get(e.action_flag, str(e.action_flag)),
                'color': ACTION_COLOR.get(e.action_flag, 'secondary'),
                'detail': e.get_change_message() or '—',
            }
            for e in recent
        ]
        return ctx


@login_required
@permission_required('lb_manager.view_lbphysical', raise_exception=True)
def lb_inventory_data(request):
    """
    AJAX endpoint — unified LBGuest + LBPhysical inventory for DataTables.

    Executes the canonical UNION SQL query directly against PostgreSQL so the
    result is identical to running the query manually in a DB client.

    Columns returned (0-indexed):
      0  type            – 'Guest' or 'Physical'
      1  device          – Hostname of the LB
      2  ci_id           – Configuration Item ID
      3  mgmt_ip         – Management IP address
      4  name            – Client company name
      5  client_code     – Company client code
      6  version         – Software version
      7  model           – Hardware/software model
      8  serial          – Serial (COALESCE: physical serial takes priority for guests)
      9  virtualizador   – Physical host device name (COALESCE: physical device for guests)
      10 environment     – DRP / PRE-PRODUCTION / PRODUCTION
      11 service         – Vendor service/support expiry date
      12 distro          – OS/distribution
      13 vendor_support  – Vendor support level
      14 datacenter      – Data-center name
      15 location        – Physical rack/room location
      16 snow_link       – ServiceNow CI link
      17 last_modified   – Timestamp of last record update
      18 monitoreo       – Monitoring integration enabled
      19 cmdb_snmp       – CMDB SNMP integration enabled
      20 cyberark        – CyberArk integration enabled
    """
    SQL = """
        SELECT
            'Guest'                                        AS type,
            lb_guest.device,
            lb_guest.ci_id,
            lb_guest.mgmt_ip,
            company.name,
            company.client_code,
            lb_guest.version,
            lb_guest.model,
            COALESCE(lb_physical.serial,  lb_guest.serial) AS serial,
            COALESCE(lb_physical.device,  lb_guest.device) AS virtualizador,
            lb_guest.environment,
            lb_physical.service,
            lb_guest.distro,
            lb_physical.vendor_support,
            datacenter.datacenter,
            lb_physical.location,
            lb_guest.snow_link,
            lb_guest.last_modified,
            lb_guest.monitoreo,
            lb_guest.cmdb_snmp,
            lb_guest.cyberark
        FROM lb_guest
        LEFT JOIN lb_physical  ON lb_guest.serial         = lb_physical.serial
        LEFT JOIN company      ON lb_guest.company_id     = company.id
        LEFT JOIN datacenter   ON lb_physical.datacenter_id = datacenter.id

        UNION

        SELECT
            'Physical'                AS type,
            lb_physical.device,
            lb_physical.ci_id,
            lb_physical.mgmt_ip,
            company.name,
            company.client_code,
            lb_physical.version,
            lb_physical.model,
            lb_physical.serial        AS serial,
            lb_physical.device        AS virtualizador,
            lb_physical.environment,
            lb_physical.service,
            lb_physical.distro,
            lb_physical.vendor_support,
            datacenter.datacenter,
            lb_physical.location,
            lb_physical.snow_link,
            lb_physical.last_modified,
            lb_physical.monitoreo,
            lb_physical.cmdb_snmp,
            lb_physical.cyberark
        FROM lb_physical
        LEFT JOIN company    ON lb_physical.company_id    = company.id
        LEFT JOIN datacenter ON lb_physical.datacenter_id = datacenter.id
    """

    # Column names aligned with the SELECT position (0-indexed)
    _INV_SORT_COLS = [
        'type', 'device', 'ci_id', 'mgmt_ip', 'name', 'client_code',
        'version', 'model', 'serial', 'virtualizador', 'environment',
        'service', 'distro', 'vendor_support', 'datacenter', 'location',
        'snow_link', 'last_modified', 'monitoreo', 'cmdb_snmp', 'cyberark',
    ]
    # Text-searchable columns (exclude booleans and dates to keep the WHERE fast)
    _INV_SEARCH_COLS = [
        'type', 'device', 'ci_id', 'mgmt_ip', 'name', 'client_code',
        'version', 'model', 'serial', 'virtualizador', 'environment',
        'distro', 'vendor_support', 'datacenter', 'location',
    ]

    draw    = int(request.GET.get('draw', 1))
    start   = int(request.GET.get('start', 0))
    length  = _dt_length(request)
    search  = request.GET.get('search[value]', '').strip().lower()
    col_idx = int(request.GET.get('order[0][column]', 1))
    col_dir = request.GET.get('order[0][dir]', 'asc')

    sort_col  = _INV_SORT_COLS[col_idx] if col_idx < len(_INV_SORT_COLS) else 'device'
    direction = 'DESC' if col_dir == 'desc' else 'ASC'

    # ── Total (before search) ─────────────────────────────────────────────────
    with connection.cursor() as cursor:
        cursor.execute(f'SELECT COUNT(*) FROM ({SQL}) AS _inv_total')
        total = cursor.fetchone()[0]

    # ── Build optional WHERE clause for full-text search ──────────────────────
    where_sql    = ''
    where_params = []
    if search:
        like = f'%{search}%'
        conditions = ' OR '.join(
            f"lower(COALESCE({c}::text, '')) LIKE %s" for c in _INV_SEARCH_COLS
        )
        where_sql    = f'WHERE {conditions}'
        where_params = [like] * len(_INV_SEARCH_COLS)

    company = request.GET.get('company', '').strip()
    if company:
        company_cond = "lower(COALESCE(client_code::text, '')) = lower(%s)"
        where_sql    = f'WHERE {company_cond}' if not where_sql else f'{where_sql} AND {company_cond}'
        where_params.append(company)

    # ── Filtered count ────────────────────────────────────────────────────────
    with connection.cursor() as cursor:
        cursor.execute(
            f'SELECT COUNT(*) FROM ({SQL}) AS _inv_filtered {where_sql}',
            where_params,
        )
        filtered = cursor.fetchone()[0]

    # ── Fetch only the requested page, sorted at DB level ────────────────────
    # sort_col comes from _INV_SORT_COLS whitelist (safe for identifier interpolation).
    # direction is hardcoded to 'ASC'/'DESC' (safe).
    # LIMIT/OFFSET use parameterized placeholders to avoid any injection risk.
    if length != -1:
        limit_sql    = 'LIMIT %s OFFSET %s'
        limit_params = [length, start]
    else:
        limit_sql    = ''
        limit_params = []
    with connection.cursor() as cursor:
        cursor.execute(
            f'SELECT * FROM ({SQL}) AS _inv {where_sql}'
            f' ORDER BY {sort_col} {direction} NULLS LAST {limit_sql}',
            where_params + limit_params,
        )
        raw_rows = cursor.fetchall()

    # Normalise: None → '-', dates/datetimes → strings, booleans → 'True'/'False'
    def _fmt(i: int, val) -> str:
        if val is None:
            return '-'
        if i == 11:
            return str(val)             # service (date)
        if i == 17:
            return str(val)[:16]        # last_modified (datetime)
        if i in (18, 19, 20):
            return 'True' if val else 'False'  # boolean flags
        return val if isinstance(val, str) else str(val)

    rows = [[_fmt(i, val) for i, val in enumerate(raw)] for raw in raw_rows]

    return JsonResponse({'draw': draw, 'recordsTotal': total, 'recordsFiltered': filtered, 'data': rows})


# ============================================================
#  AUTOCOMPLETE
# ============================================================
@login_required
@permission_required('lb_manager.view_lbphysical', raise_exception=True)
def autocomplete_data(request):
    """
    AJAX endpoint — return distinct non-empty values for a given model field.

    Used by the LBGuest and LBPhysical create/edit forms to suggest existing
    values as the user types (HTML5 <datalist> autocomplete).

    GET parameters:
        model  – 'lb_guest' or 'lb_physical'
        field  – one of the allowed fields for that model

    Returns JSON: {"values": ["val1", "val2", ...]}
    """
    ALLOWED = {
        'lb_guest': {
            'model': LBGuest,
            'fields': ['version', 'distro', 'model'],
        },
        'lb_physical': {
            'model': LBPhysical,
            'fields': ['version', 'distro', 'model', 'vendor_support', 'location'],
        },
    }

    model_key = request.GET.get('model', '')
    field     = request.GET.get('field', '')

    entry = ALLOWED.get(model_key)
    if not entry or field not in entry['fields']:
        return JsonResponse({'values': []})

    values = list(
        entry['model'].objects
            .exclude(**{f'{field}__isnull': True})
            .exclude(**{f'{field}': ''})
            .values_list(field, flat=True)
            .distinct()
            .order_by(field)[:200]
    )
    return JsonResponse({'values': values})


@login_required
@permission_required('lb_manager.view_lbphysical', raise_exception=True)
def cmdb_vs_lb_inventory(request):
    """
    Upload a CMDB CSV export and compare it against LBPhysical + LBGuest
    records for a given Company.

    GET  → show form (company selector + file upload)
    POST → process file, return Excel comparison report
    """
    companies = Company.objects.order_by('name')

    if request.method != 'POST':
        return render(request, 'lb_manager/cmdb_validation.html', {
            'companies': companies,
        })

    # ── read inputs ──────────────────────────────────────────────────────────
    company_id  = request.POST.get('company_id')
    upload_file = request.FILES.get('cmdb_file')

    if not company_id or not upload_file:
        messages.error(request, 'Debes seleccionar una empresa y subir el archivo CSV.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    _MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
    if upload_file.size > _MAX_UPLOAD_BYTES:
        messages.error(request, f'El archivo supera el límite de 10 MB ({upload_file.size // 1024 // 1024} MB).')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    try:
        company_obj = Company.objects.get(pk=company_id)
    except Company.DoesNotExist:
        messages.error(request, 'Empresa no encontrada.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    # ── parse CSV ────────────────────────────────────────────────────────────
    try:
        text    = upload_file.read().decode('utf-8-sig')  # handle BOM
        reader  = csv.DictReader(io.StringIO(text))
        cmdb_rows = list(reader)
    except (UnicodeDecodeError, csv.Error, LookupError, OSError) as exc:
        messages.error(request, f'Error al leer el archivo CSV: {exc}')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    if not cmdb_rows:
        messages.error(request, 'El archivo CSV está vacío.')
        return render(request, 'lb_manager/cmdb_validation.html', {'companies': companies})

    # ── load DB devices for the selected company (keyed by short hostname) ─────
    all_physicals = LBPhysical.objects.filter(company=company_obj).select_related('company')
    all_guests    = LBGuest.objects.filter(company=company_obj).select_related('company')

    # Primary index: short hostname (device up to first '.') → LB object
    db_by_name = {}
    for d in list(all_physicals) + list(all_guests):
        short = d.device.split('.')[0].lower()
        db_by_name[short] = d

    db_names = set(db_by_name.keys())

    # ── load configured expected values ──────────────────────────────────────
    ag_values = [
        v.lower() for v in
        CMDBFieldConfig.objects.filter(
            field_name=CMDBFieldConfig.Field.ASSIGNMENT_GROUP, active=True
        ).values_list('expected_value', flat=True)
    ]
    sg_values = [
        v.lower() for v in
        CMDBFieldConfig.objects.filter(
            field_name=CMDBFieldConfig.Field.SUPPORT_GROUP, active=True
        ).values_list('expected_value', flat=True)
    ]

    today = dj_timezone.localdate()

    # ── compare each CMDB row ─────────────────────────────────────────────────
    output_rows   = []
    cmdb_names    = set()

    for row in cmdb_rows:
        cmdb_name  = (row.get('name') or '').strip().lower()
        if cmdb_name:
            cmdb_names.add(cmdb_name)

        lb_device = db_by_name.get(cmdb_name)

        if not cmdb_name or lb_device is None:
            result = {
                'r_encontrado':       'No encontrado en DB' if cmdb_name else 'name vacío',
                'r_fqdn':             _NA,
                'r_asset_tag':        _NA,
                'r_company':          _NA,
                'r_used_for':         _NA,
                'r_ip_address':       _NA,
                'r_adm_ip_address':   _NA,
                'r_serial':           _NA,
                'r_last_discovered':  _NA,
                'r_assignment_group': _NA,
                'r_support_group':    _NA,
                'r_resumen':          'No encontrado en DB',
            }
        else:
            result = _cmdb_compare_row(row, lb_device, ag_values, sg_values, today)
            result['r_encontrado'] = 'Encontrado'

        output_rows.append({**row, **result})

    # ── devices in DB but NOT in CMDB ────────────────────────────────────────
    only_in_db = db_names - cmdb_names
    for short_name in sorted(only_in_db):
        lb = db_by_name[short_name]
        empty_cmdb = {col: '' for col in _CMDB_COLS}
        empty_cmdb['name'] = short_name
        empty_cmdb['fqdn'] = lb.device
        result = {col: _NA for col in _RESULT_COLS}
        result['r_encontrado'] = 'Solo en DB — no está en el archivo CMDB'
        result['r_resumen']    = 'Solo en DB'
        output_rows.append({**empty_cmdb, **result})

    # ── generate Excel with openpyxl ─────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN  = PatternFill('solid', fgColor='C6EFCE')
    RED    = PatternFill('solid', fgColor='FFC7CE')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')
    GREY   = PatternFill('solid', fgColor='D9D9D9')
    BLUE   = PatternFill('solid', fgColor='BDD7EE')
    HEADER_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Validación CMDB vs DB'

    # determine all columns present in the CSV
    cmdb_col_names = list(cmdb_rows[0].keys()) if cmdb_rows else _CMDB_COLS
    all_cols = list(cmdb_col_names) + _RESULT_COLS
    headers  = {col: idx + 1 for idx, col in enumerate(all_cols)}

    # ── header row ───────────────────────────────────────────────────────────
    for col_name, col_idx in headers.items():
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill      = BLUE if col_name in _RESULT_COLS else GREY
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # ── data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(output_rows, start=2):
        for col_name, col_idx in headers.items():
            value = row.get(col_name, '')
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=False)
            cell.border    = THIN_BORDER

            # colour result columns
            if col_name in _RESULT_COLS:
                if value == _OK:
                    cell.fill = GREEN
                elif value == _NA:
                    cell.fill = GREY
                elif 'Revisar' in value:
                    cell.fill = YELLOW
                elif col_name == 'r_resumen' and value == 'Todo OK':
                    cell.fill = GREEN
                else:
                    cell.fill = RED

    # ── column widths ─────────────────────────────────────────────────────────
    for col_name, col_idx in headers.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 30 if col_name in _RESULT_COLS else 22

    # ── summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumen')
    total_rows   = len(output_rows)
    found        = sum(1 for r in output_rows if r.get('r_encontrado') == 'Encontrado')
    not_found    = sum(1 for r in output_rows if 'No encontrado' in r.get('r_encontrado', ''))
    only_db      = sum(1 for r in output_rows if 'Solo en DB' in r.get('r_encontrado', ''))
    all_ok       = sum(1 for r in output_rows if r.get('r_resumen') == 'Todo OK')
    with_diffs   = sum(1 for r in output_rows if r.get('r_encontrado') == 'Encontrado'
                       and r.get('r_resumen') != 'Todo OK')

    summary = [
        ('Empresa validada',           company_obj.name),
        ('Fecha de validación',        str(today)),
        ('',                           ''),
        ('Total filas procesadas',     total_rows),
        ('Encontrados en DB',          found),
        ('No encontrados en DB',       not_found),
        ('Solo en DB (no en CMDB)',    only_db),
        ('',                           ''),
        ('Sin diferencias (Todo OK)',  all_ok),
        ('Con diferencias',            with_diffs),
    ]
    for s_row_idx, (label, value) in enumerate(summary, start=1):
        ws2.cell(row=s_row_idx, column=1, value=label).font = Font(bold=bool(label))
        ws2.cell(row=s_row_idx, column=2, value=value)
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 20

    # ── return file ───────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Strip characters that would break the Content-Disposition header value.
    safe_name = re.sub(r'[^\w\-.]', '_', company_obj.name)
    filename = f'cmdb_validacion_{safe_name}_{today}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Ansible Inventory File Editor ─────────────────────────────────────────────

def _parse_ini_hostnames(content):
    """
    Return the set of hostnames defined in an Ansible INI inventory string.

    Lines inside a [group:children] section are group names, not hosts, so
    they are skipped.  Comment lines (#) and section headers ([…]) are also
    skipped.  The first whitespace-delimited token on each remaining line is
    taken as the hostname (handles 'host ansible_host=… ansible_user=…').
    """
    hostnames = set()
    in_children = False
    for raw in content.splitlines():
        line = raw.strip()
        if not line or line.startswith('#'):
            continue
        if line.startswith('['):
            in_children = ':children' in line
            continue
        if not in_children:
            hostnames.add(line.split()[0])
    return hostnames


@login_required
@permission_required('lb_manager.view_lbguest', raise_exception=True)
def ansible_inventory(request):
    """
    Ansible Inventory File Editor.

    Displays the list of configured inventory files (AnsibleInventoryFile).
    When one is selected:
      GET  ?inv=<pk>  → reads file from disk, shows it in a textarea, and
                        lists LBGuest hosts for that environment that are not
                        yet present in the file.
      POST ?inv=<pk>  → writes the textarea content back to the file.
    """
    inv_files     = AnsibleInventoryFile.objects.order_by('environment', 'name')
    inv_id        = request.GET.get('inv') or request.POST.get('inv')
    selected      = None
    file_content  = ''
    file_error    = None
    missing_hosts = []
    orphan_hosts  = []

    if inv_id:
        try:
            selected = AnsibleInventoryFile.objects.get(pk=inv_id)
        except AnsibleInventoryFile.DoesNotExist:
            messages.error(request, 'Inventario no encontrado.')

    if selected:
        if request.method == 'POST':
            if not request.user.has_perm('lb_manager.change_lbguest'):
                messages.error(request, 'No tienes permiso para modificar archivos de inventario.')
                file_content = request.POST.get('content', '')
            else:
                file_content = request.POST.get('content', '')
                try:
                    with open(selected.file_path, 'w', encoding='utf-8') as fh:
                        fh.write(file_content)
                    messages.success(request, f'Inventario guardado: {selected.file_path}')
                except OSError as exc:
                    messages.error(request, f'Error al guardar el archivo: {exc}')
        else:
            try:
                with open(selected.file_path, 'r', encoding='utf-8') as fh:
                    file_content = fh.read()
            except FileNotFoundError:
                file_error = f'Archivo no encontrado: {selected.file_path}'
            except OSError as exc:
                file_error = f'Error al leer el archivo: {exc}'

        # Cross-check inventory file vs DB
        if file_content:
            inventory_hosts = _parse_ini_hostnames(file_content)

            # Hosts in DB (correct environment) that are absent from the file
            db_hosts = list(
                LBGuest.objects
                .filter(environment__iexact=selected.environment)
                .values_list('device', flat=True)
                .order_by('device')
            )
            missing_hosts = [h for h in db_hosts if h not in inventory_hosts]

            # Hosts in the file that are absent from DB or have wrong environment
            db_all = {
                row['device']: row['environment']
                for row in LBGuest.objects.values('device', 'environment')
            }
            def _orphan_reason(h: str) -> str | None:
                if h not in db_all:
                    return 'No existe en la DB'
                if db_all[h].upper() != selected.environment.upper():
                    return f'En DB como {db_all[h]}'
                return None

            orphan_hosts = [
                {'host': h, 'reason': r}
                for h in sorted(inventory_hosts)
                if (r := _orphan_reason(h)) is not None
            ]

    return render(request, 'lb_manager/ansible_inventory.html', {
        'inv_files':     inv_files,
        'selected':      selected,
        'file_content':  file_content,
        'file_error':    file_error,
        'missing_hosts': missing_hosts,
        'orphan_hosts':  orphan_hosts,
    })
